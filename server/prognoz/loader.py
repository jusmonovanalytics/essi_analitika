"""Excel savdo hisobotlarini o'qish va bazaga yuklash.

Ikkala manba ham bir xil 16 ustunli formatda keladi. Farqi — kunlar soni:
    fakt savdo    — bitta kun (talab: mijoz nima so'ragan)
    yakuniy savdo — bir necha kun (ombor kesimidan keyin nima yetkazilgan)

MUHIM: yuklash prognozni QAYTA HISOBLAMAYDI. Agregatlar yangilanadi, prognoz
esa eski holicha qoladi — toki foydalanuvchi qo'lda qayta hisoblamaguncha.

Excel o'qish — bloklovchi amal, shuning uchun u threadpool'da bajariladi
(`asyncio.to_thread`), event loop to'xtab qolmasin.
"""
import io
import os

import openpyxl

import db as core

# Excel ustun indekslari (0-ustun — tartib raqami)
C_DATE, C_AGENT, C_ORDERER, C_COURIER = 1, 2, 3, 4
C_ZONE, C_SHOPTYPE, C_SHOP, C_SHOPNO = 5, 6, 7, 8
C_PTYPE, C_PRODUCT, C_ORDERNO = 9, 10, 11
C_PAY, C_DISC, C_QTY, C_AMOUNT = 12, 13, 14, 15

EXPECTED_HEADER = (
    "По дням", "Ответственный агент магазина", "Заказ оформил", "Доставщик",
    "Зона", "Тип магазина", "Магазин", "№ Магазина", "Тип продукта", "Продукт",
    "№ заказа", "Тип оплаты", "Процент скидки", "Колв.продуктов БВ", "Общ.сумма БВ",
)

FAKT_COLS = ("sale_date", "agent", "orderer", "courier", "zone", "shop_type",
             "shop_name", "shop_no", "product_type", "product", "order_no",
             "pay_type", "discount_pct", "qty", "amount", "source_file")

YAK_COLS = ("sale_date", "week_range", "agent", "orderer", "courier", "zone",
            "shop_type", "shop_name", "shop_no", "product_type", "product",
            "order_no", "pay_type", "discount_pct", "qty", "amount", "source_file")


class BadFile(Exception):
    """Fayl kutilgan formatda emas."""


def _norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


def _to_num(v):
    """Sonli qiymat; 'Не указано' kabi matn bo'lsa — None."""
    if v is None:
        return None
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def read_workbook(data: bytes) -> list[dict]:
    """Excel'ni o'qiydi va tekshiradi. BLOKLOVCHI — to_thread da chaqiring."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:                                       # noqa: BLE001
        raise BadFile(f"Excel fayl sifatida ochib bo'lmadi: {e}")

    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        raise BadFile("Fayl bo'sh")

    got = tuple(_norm(h) for h in header[1:16])
    if got != EXPECTED_HEADER:
        missing = [h for h in EXPECTED_HEADER if h not in got]
        raise BadFile(
            "Ustunlar mos kelmadi. Kutilgan: savdo hisoboti (16 ustun). "
            + (f"Topilmadi: {', '.join(missing[:4])}" if missing
               else "Ustunlar tartibi boshqacha.")
        )

    rows = []
    for i, r in enumerate(it, start=2):
        if r[C_DATE] is None or r[C_PRODUCT] is None:
            continue
        try:
            rows.append({
                "date": r[C_DATE].date() if hasattr(r[C_DATE], "date") else r[C_DATE],
                "agent": _norm(r[C_AGENT]),
                "orderer": _norm(r[C_ORDERER]),
                "courier": _norm(r[C_COURIER]),
                "zone": _norm(r[C_ZONE]),
                "shop_type": _norm(r[C_SHOPTYPE]),
                "shop_name": _norm(r[C_SHOP]),
                "shop_no": _to_int(r[C_SHOPNO]),
                "product_type": _norm(r[C_PTYPE]),
                "product": _norm(r[C_PRODUCT]),
                "order_no": _to_int(r[C_ORDERNO]),
                "pay_type": _norm(r[C_PAY]) or "unknown",
                "discount_pct": _to_num(r[C_DISC]),
                "qty": _to_num(r[C_QTY]) or 0.0,
                "amount": _to_num(r[C_AMOUNT]) or 0.0,
            })
        except (TypeError, ValueError) as e:
            raise BadFile(f"{i}-qatorda noto'g'ri qiymat: {e}")
    wb.close()

    if not rows:
        raise BadFile("Faylda ma'lumot qatori yo'q")
    return rows


async def _copy(cur, table: str, cols: tuple[str, ...], records: list[tuple]):
    async with cur.copy(
        f"COPY {table} ({', '.join(cols)}) FROM STDIN"
    ) as cp:
        for rec in records:
            await cp.write_row(rec)


async def load_fakt(source_file: str, rows: list[dict], replace: bool = False) -> dict:
    """Fakt savdo — bitta kunlik hisobot.

    Dublikat himoyasi SANA bo'yicha (fayl nomi bo'yicha emas — nomlar xilma-xil:
    `report (3).xlsx`, `5465132132151515.xlsx`).
    """
    dates = {r["date"] for r in rows}
    if len(dates) > 1:
        raise BadFile(
            f"Faylda {len(dates)} xil sana bor ({min(dates)} … {max(dates)}). "
            "Fakt savdo bitta kunga tegishli bo'lishi kerak. "
            "Bu yakuniy savdo fayli bo'lsa — manbani «Yakuniy savdo» qilib tanlang."
        )
    sale_date = next(iter(dates))

    pool = await core.get_pool()
    async with pool.connection() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT count(*), min(source_file) FROM fakt_savdo WHERE sale_date = %s",
                (sale_date,))
            n, old_file = await cur.fetchone()

            if n and not replace:
                return {"status": "skipped", "sale_date": sale_date, "rows": 0,
                        "reason": f"{sale_date} allaqachon yuklangan (fayl: {old_file})"}
            if n and replace:
                await cur.execute("DELETE FROM fakt_savdo WHERE sale_date = %s",
                                  (sale_date,))

            recs = [(
                r["date"], r["agent"], r["orderer"], r["courier"], r["zone"],
                r["shop_type"], r["shop_name"], r["shop_no"], r["product_type"],
                r["product"], r["order_no"], r["pay_type"], r["discount_pct"],
                int(r["qty"]), r["amount"], source_file,
            ) for r in rows]
            await _copy(cur, "fakt_savdo", FAKT_COLS, recs)

    return {"status": "replaced" if n else "loaded", "sale_date": sale_date,
            "rows": len(rows), "qty": sum(r["qty"] for r in rows),
            "amount": sum(r["amount"] for r in rows)}


async def load_yakuniy(source_file: str, rows: list[dict], replace: bool = False) -> dict:
    """Yakuniy savdo — haftalik fayl, bir necha kun bo'lishi mumkin.

    Dublikat himoyasi FAYL NOMI bo'yicha.
    """
    week = os.path.splitext(source_file)[0]

    pool = await core.get_pool()
    async with pool.connection() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT count(*) FROM yakuniy_savdo WHERE source_file = %s",
                (source_file,))
            n = (await cur.fetchone())[0]

            if n and not replace:
                return {"status": "skipped", "rows": 0,
                        "reason": f"«{source_file}» allaqachon yuklangan"}
            if n and replace:
                await cur.execute("DELETE FROM yakuniy_savdo WHERE source_file = %s",
                                  (source_file,))

            recs = [(
                r["date"], week, r["agent"], r["orderer"], r["courier"], r["zone"],
                r["shop_type"], r["shop_name"], r["shop_no"], r["product_type"],
                r["product"], r["order_no"], r["pay_type"], r["discount_pct"],
                r["qty"], r["amount"], source_file,
            ) for r in rows]
            await _copy(cur, "yakuniy_savdo", YAK_COLS, recs)

    ds = sorted({r["date"] for r in rows})
    return {"status": "replaced" if n else "loaded", "rows": len(rows),
            "kunlar": len(ds), "dan": ds[0], "gacha": ds[-1],
            "qty": sum(r["qty"] for r in rows)}
