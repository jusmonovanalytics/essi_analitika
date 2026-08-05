"""Savdo prognozi — arxiv, qo'lda tahrir, Excel yuklash/o'chirish, eksport.

router.py bilan bitta APIRouter'ga yig'iladi.
"""
import asyncio
import io
import json
import os
import re
from collections import defaultdict
from datetime import date, timedelta

from fastapi import Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from . import db
from .auth import admin
from .loader import BadFile, load_fakt, load_yakuniy, read_workbook
from .router import router, _run


# Bazani o'zgartiradigan har bir endpoint shu qo'riqchidan o'tadi.
# O'qish (ko'rish, filtrlash, Excel yuklab olish) parolsiz qoladi.
ADMIN = [Depends(admin)]


@router.post("/parol")
async def parol_tekshir(_: None = Depends(admin)):
    """Parolni tekshirish — interfeys uni saqlashdan oldin chaqiradi."""
    return {"ok": True}


# ═══════════════════════════════════════════════════════════════════════════
#   ARXIV — har bir hisob va tahrir saqlanadi, hech qachon o'chmaydi
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/arxiv")
async def arxiv(limit: int = Query(50, ge=1, le=200)):
    rows = await db.q("SELECT * FROM v_arxiv LIMIT %s", (limit,))
    return [{
        "run_id": r["run_id"], "created_at": r["created_at"].isoformat(),
        "faol": r["faol"], "data_last_day": r["data_last_day"].isoformat(),
        "gorizont": r["gorizont"], "ustama": r["ustama"], "zaxira": float(r["kalibr"]),
        "mahsulot": r["n_mahsulot"], "jami": float(r["jami_qty"]),
        "dan": r["dan"].isoformat(), "gacha": r["gacha"].isoformat(),
        "fakt_kunlar": r["fakt_kunlar"], "fakt_fayllar": r["fakt_fayllar"],
        "yak_kunlar": r["yak_kunlar"],
        "qolda": r["qolda"], "asos_run": r["asos_run"],
        "ozgartirilgan": r["ozgartirilgan_qator"],
        "farq": float(r["farq_oldingidan"]) if r["farq_oldingidan"] else None,
        "izoh": r["izoh"],
    } for r in rows]


@router.get("/arxiv/{run_id}")
async def arxiv_bitta(run_id: int):
    """Arxivdagi bitta rejaning to'liq tarkibi."""
    r = await db.q("SELECT * FROM v_arxiv WHERE run_id = %s", (run_id,), one=True)
    if not r:
        raise HTTPException(404, f"run_id={run_id} arxivda yo'q")

    items = await db.q("""
        SELECT p.product_id, p.name, p.product_type,
               sum(d.qty) total, sum(d.qty_past) lo, sum(d.qty_yuqori) hi,
               sum(d.qty) - sum(COALESCE(d.qty_model, d.qty)) AS qolda_farq
        FROM reja_daily d JOIN products p USING (product_id)
        WHERE d.run_id = %s
        GROUP BY 1, 2, 3 ORDER BY total DESC
    """, (run_id,))

    return {
        "run": {
            "run_id": r["run_id"], "created_at": r["created_at"].isoformat(),
            "faol": r["faol"], "data_last_day": r["data_last_day"].isoformat(),
            "jami": float(r["jami_qty"]), "zaxira": float(r["kalibr"]),
            "gorizont": r["gorizont"], "ustama": r["ustama"],
            "mahsulot": r["n_mahsulot"],
            "dan": r["dan"].isoformat(), "gacha": r["gacha"].isoformat(),
            "fakt_kunlar": r["fakt_kunlar"], "fakt_fayllar": r["fakt_fayllar"],
            "yak_kunlar": r["yak_kunlar"],
            "farq": float(r["farq_oldingidan"]) if r["farq_oldingidan"] else None,
            "qolda": r["qolda"], "asos_run": r["asos_run"],
            "ozgartirilgan": r["ozgartirilgan_qator"], "izoh": r["izoh"],
        },
        "items": [{
            "product_id": i["product_id"], "name": i["name"], "type": i["product_type"],
            "total": round(float(i["total"])), "lo": round(float(i["lo"])),
            "hi": round(float(i["hi"])),
            "qolda_farq": round(float(i["qolda_farq"] or 0)),
        } for i in items],
    }


@router.get("/arxiv/{run_id}/pivot")
async def arxiv_pivot(run_id: int, round_to: int = Query(50, ge=1, le=500)):
    """Arxivdagi rejaning TO'LIQ kunlik jadvali (mahsulot × kun).

    Joriy reja bilan bir xil ko'rinish — lekin o'sha paytdagi holat.
    `model` — modelning asl qiymati; `edited` — qo'lda o'zgartirilganmi.
    """
    run = await db.q("SELECT * FROM v_arxiv WHERE run_id = %s", (run_id,), one=True)
    if not run:
        raise HTTPException(404, f"run_id={run_id} arxivda yo'q")

    rows = await db.q("""
        SELECT p.product_id, p.name AS product, p.product_type,
               d.target_date, d.qty, COALESCE(d.qty_model, d.qty) AS qty_model,
               (d.qty <> COALESCE(d.qty_model, d.qty)) AS ozgartirilgan
        FROM reja_daily d JOIN products p USING (product_id)
        WHERE d.run_id = %s
        ORDER BY p.product_type, p.name, d.target_date
    """, (run_id,))
    if not rows:
        raise HTTPException(404, "Reja bo'sh")

    cells, model, edited, meta, dates = {}, {}, {}, {}, set()
    for r in rows:
        pid, td = r["product_id"], r["target_date"]
        cells.setdefault(pid, {})[td] = float(r["qty"])
        model.setdefault(pid, {})[td] = float(r["qty_model"])
        edited.setdefault(pid, {})[td] = bool(r["ozgartirilgan"])
        meta[pid] = (r["product"], r["product_type"])
        dates.add(td)

    def rnd(v):
        return int(round(v / round_to) * round_to)

    d0, d1 = min(dates), max(dates)
    col_dates, d = [], d0
    while d <= d1:
        col_dates.append(d)
        d += timedelta(days=1)
    columns = [{"date": c.isoformat(), "dow": c.isoweekday(), "reja": c in dates}
               for c in col_dates]

    out = []
    for pid, (name, pt) in meta.items():
        vals = [rnd(cells[pid][c]) if c in cells[pid] else None for c in col_dates]
        out.append({
            "product_id": pid, "name": name, "type": pt, "values": vals,
            "model": [rnd(model[pid][c]) if c in model[pid] else None for c in col_dates],
            "edited": [bool(edited[pid].get(c)) if c in cells[pid] else False
                       for c in col_dates],
            "total": sum(v for v in vals if v),
        })
    out.sort(key=lambda r: r["name"])

    totals = [sum(r["values"][i] or 0 for r in out) if c["reja"] else None
              for i, c in enumerate(columns)]

    return {
        "run_id": run_id, "round_to": round_to,
        "columns": columns, "rows": out, "totals": totals,
        "jami": sum(r["total"] for r in out),
        "faol": run["faol"], "qolda": run["qolda"],
        "ozgartirilgan": run["ozgartirilgan_qator"],
        "tahrirlanadi": False,          # arxiv o'zgarmas
    }


@router.get("/arxiv/{run_id}/eksport")
async def arxiv_eksport(run_id: int):
    """Arxivdagi rejani Excel'ga chiqaradi (потребность formatida)."""
    run = await db.q("SELECT * FROM v_arxiv WHERE run_id = %s", (run_id,), one=True)
    if not run:
        raise HTTPException(404, f"run_id={run_id} arxivda yo'q")

    rows = await db.q("""
        SELECT p.name AS product, d.target_date, d.qty
        FROM reja_daily d JOIN products p USING (product_id)
        WHERE d.run_id = %s ORDER BY p.name, d.target_date
    """, (run_id,))
    if not rows:
        raise HTTPException(404, "Reja bo'sh")

    data = await asyncio.to_thread(_excel, rows, run)
    nom = f"potrebnost_{run['dan']}_{run['gacha']}_run{run_id}"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom}.xlsx"'},
    )


@router.post("/arxiv/{run_id}/faollashtir", dependencies=ADMIN)
async def faollashtir(run_id: int):
    """Arxivdagi eski rejaga qaytish. Joriy reja arxivda qoladi."""
    try:
        msg = await db.x("SELECT fn_reja_faollashtir(%s)", (run_id,))
    except Exception as e:                                       # noqa: BLE001
        raise HTTPException(400, str(e).split("CONTEXT")[0].strip())
    return {"ok": True, "message": msg}


@router.post("/hisobla", dependencies=ADMIN)
async def hisobla(
    gorizont: int = Query(12, ge=6, le=30),
    ustama: bool = Query(True),
    zaxira: float = Query(1.03, ge=0.5, le=1.5),
    izoh: str = Query(None),
    boshlanish: date = Query(None, description="Reja qaysi kundan boshlanadi. "
                                               "Bo'sh = ma'lumotdan keyingi ish kuni."),
):
    """QAYTA HISOBLASH — yagona nuqta. Avtomatik chaqirilmaydi.

    Eski reja arxivda qoladi, yangisi uning yoniga qo'shiladi.
    """
    eski = await db.joriy_run()
    try:
        # Tiplarni ANIQ belgilaymiz. psycopg Python float ni `double precision`
        # qilib yuboradi, PostgreSQL esa uni `numeric` ga o'zi o'girmaydi —
        # castsiz fn_reja_saqla(smallint, boolean, double precision, unknown)
        # topilmay, so'rov 500 bilan yiqiladi.
        run_id = await db.x(
            "SELECT fn_reja_saqla(%s::int, %s::boolean, %s::numeric, %s::text, %s::date)",
            (gorizont, ustama, zaxira, izoh, boshlanish))
    except Exception as e:                                       # noqa: BLE001
        raise HTTPException(500, str(e).split("CONTEXT")[0].strip())

    yangi = await db.q("SELECT * FROM reja_runs WHERE run_id = %s", (run_id,), one=True)
    return {
        "ok": True, "run_id": run_id,
        "jami": float(yangi["jami_qty"]), "mahsulot": yangi["n_mahsulot"],
        "dan": yangi["dan"].isoformat(), "gacha": yangi["gacha"].isoformat(),
        "arxivlandi": {"run_id": eski["run_id"], "jami": float(eski["jami_qty"])}
                      if eski else None,
        "farq": (float(yangi["jami_qty"]) - float(eski["jami_qty"])) if eski else None,
    }


@router.post("/tahrir", dependencies=ADMIN)
async def tahrir(payload: dict = Body(...)):
    """Rejani QO'LDA tahrirlash.

    Eski reja O'ZGARMAYDI — undan yangi versiya yaratiladi. Modelning asl
    qiymati saqlanadi, shunda nimani o'zgartirganingiz doim ko'rinadi.

    body: {"changes": [{"product_id":1,"target_date":"2026-07-13","qty":500}], "izoh": "…"}
    """
    changes = payload.get("changes") or []
    if not changes:
        raise HTTPException(400, "O'zgarish yuborilmadi")
    for ch in changes:
        if not all(k in ch for k in ("product_id", "target_date", "qty")):
            raise HTTPException(
                400, "Har bir o'zgarishda product_id, target_date, qty bo'lishi kerak")
        if float(ch["qty"]) < 0:
            raise HTTPException(400, "Manfiy miqdor bo'lishi mumkin emas")

    try:
        run_id = await db.x(
            "SELECT fn_reja_qolda(%s::jsonb, %s, %s)",
            (json.dumps(changes), payload.get("asos"), payload.get("izoh")),
        )
    except Exception as e:                                       # noqa: BLE001
        raise HTTPException(400, str(e).split("CONTEXT")[0].strip())

    r = await db.q("SELECT * FROM v_arxiv WHERE run_id = %s", (run_id,), one=True)
    return {
        "ok": True, "run_id": run_id, "asos_run": r["asos_run"],
        "jami": float(r["jami_qty"]),
        "ozgartirilgan": r["ozgartirilgan_qator"],
        "farq": float(r["farq_oldingidan"]) if r["farq_oldingidan"] else 0,
    }


# ═══════════════════════════════════════════════════════════════════════════
#   MA'LUMOT — Excel yuklash va o'chirish
# ═══════════════════════════════════════════════════════════════════════════

@router.get("/holat")
async def holat():
    """Ikkala manbaning holati."""
    f = await db.q("""
        SELECT count(*) rows, sum(qty) qty, sum(amount) amount,
               min(sale_date) d0, max(sale_date) d1, count(DISTINCT sale_date) days,
               count(DISTINCT order_no) orders, count(DISTINCT source_file) files
        FROM fakt_savdo
    """, one=True)
    y = await db.q("""
        SELECT count(*) rows, sum(qty) qty, min(sale_date) d0, max(sale_date) d1,
               count(DISTINCT sale_date) days, count(DISTINCT source_file) files
        FROM yakuniy_savdo
    """, one=True)
    d = await db.q("""
        SELECT (SELECT count(*) FROM products)                  products,
               (SELECT count(DISTINCT shop_no) FROM fakt_savdo)  shops,
               (SELECT count(DISTINCT zone) FROM fakt_savdo)     zones
    """, one=True)

    return {
        "fakt": {
            "rows": f["rows"], "qty": int(f["qty"] or 0),
            "amount": float(f["amount"] or 0), "orders": f["orders"],
            "kunlar": f["days"], "fayllar": f["files"],
            "dan": f["d0"].isoformat() if f["d0"] else None,
            "gacha": f["d1"].isoformat() if f["d1"] else None,
        },
        "yakuniy": {
            "rows": y["rows"], "qty": int(y["qty"] or 0),
            "kunlar": y["days"], "fayllar": y["files"],
            "dan": y["d0"].isoformat() if y["d0"] else None,
            "gacha": y["d1"].isoformat() if y["d1"] else None,
        },
        "olchov": dict(d),
    }


OTDEL_VALUES = """
    ('Диллеры','Диллеры'), ('Магазин','Розница'), ('VIP','Розница'),
    ('OSDO','Розница'), ('A - маркет','Розница'), ('B - маркет','Розница'),
    ('С - маркет','Розница'), ('Розница','Розница'), ('Халк ретейл','Сеть'),
    ('Корзинка','Сеть'), ('Сеть','Сеть'), ('Сырная лавка','Сеть'),
    ('Макро','Сеть'), ('Урбант ретейл','Сеть'), ('Би-1','Сеть'),
    ('Магнум ретейл','Сеть'), ('Ассорти','Сеть'), ('Хавас','Хавас'),
    ('Бюджетная орг.','Хорика'), ('Школа / Садик','Хорика'),
    ('Гостиница','Хорика'), ('Кафе / Ресторан','Хорика'),
    ('Амирал Ритейл','Сеть'), ('Доставщики','Хорика'),
    ('Наманган','Диллеры'), ('Самарканд','Диллеры'), ('Бухоро','Диллеры')
"""


@router.get("/kesim-tahlili")
async def kesim_tahlili(oy: str = Query(None, description="YYYY-MM")):
    """Fakt va yakuniy savdo summasining kesilgan/qo'shilgan farqi."""
    if oy is not None and not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", oy):
        raise HTTPException(422, "oy YYYY-MM formatida bo'lishi kerak")

    months = await db.q("""
        SELECT to_char(date_trunc('month', sale_date), 'YYYY-MM') AS oy
        FROM (SELECT sale_date FROM fakt_savdo INTERSECT
              SELECT sale_date FROM yakuniy_savdo) x
        GROUP BY 1 ORDER BY 1 DESC
    """)
    available = [r["oy"] for r in months]
    selected = oy or (available[0] if available else None)
    if not selected:
        return {"oy": None, "oylar": [], "summary": {}, "kunlar": [],
                "tovarlar": [], "otdellar": []}

    start = selected + "-01"
    rows = await db.q(f"""
        WITH xarita(shop_type, otdel) AS (VALUES {OTDEL_VALUES}),
        f AS (
            SELECT sale_date, order_no, product, COALESCE(x.otdel, 'Boshqa') AS otdel,
                   sum(qty)::numeric AS qty, sum(amount)::numeric AS amount
            FROM fakt_savdo s LEFT JOIN xarita x ON btrim(s.shop_type) = x.shop_type
            WHERE s.sale_date >= %s::date AND s.sale_date < (%s::date + interval '1 month')
            GROUP BY 1,2,3,4
        ), y AS (
            SELECT sale_date, order_no, product, COALESCE(x.otdel, 'Boshqa') AS otdel,
                   sum(qty)::numeric AS qty, sum(amount)::numeric AS amount
            FROM yakuniy_savdo s LEFT JOIN xarita x ON btrim(s.shop_type) = x.shop_type
            WHERE s.sale_date >= %s::date AND s.sale_date < (%s::date + interval '1 month')
            GROUP BY 1,2,3,4
        )
        SELECT COALESCE(f.sale_date,y.sale_date) AS sana,
               COALESCE(f.order_no,y.order_no) AS order_no,
               COALESCE(f.product,y.product) AS product,
               COALESCE(f.otdel,y.otdel) AS otdel,
               COALESCE(f.qty,0) AS fakt_qty, COALESCE(y.qty,0) AS yak_qty,
               COALESCE(f.amount,0) AS fakt_summa, COALESCE(y.amount,0) AS yak_summa
        FROM f FULL JOIN y ON f.sale_date = y.sale_date
          AND f.order_no IS NOT DISTINCT FROM y.order_no
          AND f.product = y.product AND f.otdel = y.otdel
        ORDER BY 1,2,3,4
    """, (start, start, start, start))

    def empty():
        return {"fakt_summa": 0.0, "yak_summa": 0.0, "kesilgan_summa": 0.0,
                "qoshilgan_summa": 0.0, "fakt_qty": 0.0, "yak_qty": 0.0,
                "kesilgan_qty": 0.0, "qoshilgan_qty": 0.0}

    summary = empty()
    kunlar, tovarlar, otdellar = defaultdict(empty), defaultdict(empty), defaultdict(empty)

    def add(dst, row):
        fq, yq = float(row["fakt_qty"]), float(row["yak_qty"])
        fs, ys = float(row["fakt_summa"]), float(row["yak_summa"])
        dst["fakt_summa"] += fs; dst["yak_summa"] += ys
        dst["kesilgan_summa"] += max(fs - ys, 0)
        dst["qoshilgan_summa"] += max(ys - fs, 0)
        dst["fakt_qty"] += fq; dst["yak_qty"] += yq
        dst["kesilgan_qty"] += max(fq - yq, 0)
        dst["qoshilgan_qty"] += max(yq - fq, 0)

    for row in rows:
        add(summary, row)
        add(kunlar[row["sana"].isoformat()], row)
        add(tovarlar[row["product"]], row)
        add(otdellar[row["otdel"]], row)

    def packed(data, key):
        return [{key: name, **{k: round(v, 2) for k, v in values.items()}}
                for name, values in data.items()]

    return {
        "oy": selected, "oylar": available,
        "summary": {k: round(v, 2) for k, v in summary.items()},
        "kunlar": sorted(packed(kunlar, "sana"), key=lambda x: x["sana"]),
        "tovarlar": sorted(packed(tovarlar, "tovar"),
                           key=lambda x: x["kesilgan_summa"], reverse=True),
        "otdellar": sorted(packed(otdellar, "otdel"),
                           key=lambda x: x["kesilgan_summa"], reverse=True),
    }


DRILL_DIMENSIONS = {
    "sana": "sale_date::text", "otdel": "otdel", "shop_type": "shop_type", "zone": "zone",
    "agent": "agent", "orderer": "orderer", "courier": "courier",
    "shop": "shop", "pay_type": "pay_type", "product_type": "product_type",
    "product": "product", "order_no": "order_no::text",
}


@router.get("/kesim-tahlili/erkin")
async def kesim_erkin(oy: str = Query(...), group_by: str = Query(...),
                      filters: str = Query("{}")):
    """11 parametrdan istalgan tartibda erkin drill-down."""
    try:
        start = date.fromisoformat(oy + "-01")
    except ValueError:
        raise HTTPException(422, "oy YYYY-MM formatida bo'lishi kerak")
    if group_by not in DRILL_DIMENSIONS:
        raise HTTPException(422, "group_by qo'llab-quvvatlanmaydi")
    try:
        selected = json.loads(filters)
    except json.JSONDecodeError:
        raise HTTPException(422, "filters JSON obyekt bo'lishi kerak")
    if not isinstance(selected, dict) or any(k not in DRILL_DIMENSIONS for k in selected):
        raise HTTPException(422, "filters ichida noma'lum parametr bor")

    where, params = [], [start, start, start, start]
    for key, value in selected.items():
        where.append(f"COALESCE(({DRILL_DIMENSIONS[key]})::text, 'Noma''lum') = %s")
        params.append(str(value))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    group_expr = DRILL_DIMENSIONS[group_by]

    rows = await db.q(f"""
        WITH xarita(shop_type, otdel) AS (VALUES {OTDEL_VALUES}),
        f AS (
            SELECT sale_date, order_no, product, COALESCE(x.otdel,'Boshqa') otdel,
                   max(s.shop_type) shop_type, max(zone) zone, max(agent) agent,
                   max(orderer) orderer, max(courier) courier,
                   max(shop_name) shop_name, max(shop_no) shop_no,
                   max(pay_type) pay_type, max(product_type) product_type,
                   sum(qty)::numeric qty, sum(amount)::numeric amount
            FROM fakt_savdo s LEFT JOIN xarita x ON btrim(s.shop_type)=x.shop_type
            WHERE sale_date >= %s::date AND sale_date < (%s::date + interval '1 month') GROUP BY 1,2,3,4
        ), y AS (
            SELECT sale_date, order_no, product, COALESCE(x.otdel,'Boshqa') otdel,
                   max(s.shop_type) shop_type, max(zone) zone, max(agent) agent,
                   max(orderer) orderer, max(courier) courier,
                   max(shop_name) shop_name, max(shop_no) shop_no,
                   max(pay_type) pay_type, max(product_type) product_type,
                   sum(qty)::numeric qty, sum(amount)::numeric amount
            FROM yakuniy_savdo s LEFT JOIN xarita x ON btrim(s.shop_type)=x.shop_type
            WHERE sale_date >= %s::date AND sale_date < (%s::date + interval '1 month') GROUP BY 1,2,3,4
        ), d AS (
            SELECT COALESCE(f.sale_date,y.sale_date) sale_date,
                   COALESCE(f.order_no,y.order_no) order_no,
                   COALESCE(f.product,y.product) product,
                   COALESCE(f.otdel,y.otdel) otdel,
                   COALESCE(f.shop_type,y.shop_type) shop_type,
                   COALESCE(f.zone,y.zone) zone, COALESCE(f.agent,y.agent) agent,
                   COALESCE(f.orderer,y.orderer) orderer,
                   COALESCE(f.courier,y.courier) courier,
                   COALESCE(f.pay_type,y.pay_type) pay_type,
                   COALESCE(f.product_type,y.product_type) product_type,
                   COALESCE(f.shop_no::text || ' — ' || f.shop_name,
                            y.shop_no::text || ' — ' || y.shop_name,
                            f.shop_name,y.shop_name,'Noma''lum') shop,
                   COALESCE(f.qty,0) fakt_qty, COALESCE(y.qty,0) yak_qty,
                   COALESCE(f.amount,0) fakt_summa, COALESCE(y.amount,0) yak_summa
            FROM f FULL JOIN y ON f.sale_date=y.sale_date
              AND f.order_no IS NOT DISTINCT FROM y.order_no
              AND f.product=y.product AND f.otdel=y.otdel
        )
        SELECT COALESCE(({group_expr})::text,'Noma''lum') value,
               sum(fakt_summa) fakt_summa, sum(yak_summa) yak_summa,
               sum(GREATEST(fakt_summa-yak_summa,0)) kesilgan_summa,
               sum(GREATEST(yak_summa-fakt_summa,0)) qoshilgan_summa,
               sum(fakt_qty) fakt_qty, sum(yak_qty) yak_qty,
               sum(GREATEST(fakt_qty-yak_qty,0)) kesilgan_qty,
               sum(GREATEST(yak_qty-fakt_qty,0)) qoshilgan_qty,
               count(DISTINCT order_no) buyurtmalar,
               count(DISTINCT product) mahsulotlar
        FROM d {where_sql} GROUP BY 1 ORDER BY kesilgan_summa DESC, value
    """, params)
    return {"oy": oy, "group_by": group_by, "filters": selected,
            "items": [{k: (round(float(v), 2) if k not in
                       ("value", "buyurtmalar", "mahsulotlar") else v)
                       for k, v in row.items()} for row in rows]}


@router.get("/kesim-tahlili/tafsilot")
async def kesim_tafsilot(oy: str = Query(...), sana: str = Query(None), otdel: str = Query(None),
                         order_no: int = Query(None)):
    """Kun -> otdel -> buyurtma -> mahsulot drill-down ma'lumotlari."""
    try:
        start = date.fromisoformat(oy + "-01")
        if sana:
            date.fromisoformat(sana)
    except ValueError:
        raise HTTPException(422, "oy yoki sana formati noto'g'ri")
    date_filter = "s.sale_date = %s::date" if sana else \
                  "s.sale_date >= %s::date AND s.sale_date < (%s::date + interval '1 month')"
    otdel_filter = "AND COALESCE(x.otdel, 'Boshqa') = %s" if otdel else ""
    order_filter = "AND s.order_no = %s" if order_no is not None else ""
    params_f = ([sana] if sana else [start, start]) + ([otdel] if otdel else []) + ([order_no] if order_no is not None else [])
    params = params_f + params_f
    rows = await db.q(f"""
        WITH xarita(shop_type, otdel) AS (VALUES {OTDEL_VALUES}),
        f AS (
            SELECT sale_date, order_no, product, COALESCE(x.otdel, 'Boshqa') AS otdel,
                   max(product_type) product_type, max(agent) agent,
                   max(orderer) orderer, max(courier) courier, max(zone) zone,
                   max(s.shop_type) shop_type, max(shop_name) shop_name,
                   max(shop_no) shop_no, max(pay_type) pay_type,
                   max(discount_pct) discount_pct,
                   sum(qty)::numeric qty, sum(amount)::numeric amount
            FROM fakt_savdo s LEFT JOIN xarita x ON btrim(s.shop_type) = x.shop_type
            WHERE {date_filter} {otdel_filter} {order_filter}
            GROUP BY 1,2,3,4
        ), y AS (
            SELECT sale_date, order_no, product, COALESCE(x.otdel, 'Boshqa') AS otdel,
                   max(product_type) product_type, max(agent) agent,
                   max(orderer) orderer, max(courier) courier, max(zone) zone,
                   max(s.shop_type) shop_type, max(shop_name) shop_name,
                   max(shop_no) shop_no, max(pay_type) pay_type,
                   max(discount_pct) discount_pct,
                   sum(qty)::numeric qty, sum(amount)::numeric amount
            FROM yakuniy_savdo s LEFT JOIN xarita x ON btrim(s.shop_type) = x.shop_type
            WHERE {date_filter} {otdel_filter} {order_filter}
            GROUP BY 1,2,3,4
        )
        SELECT COALESCE(f.order_no,y.order_no) order_no,
               COALESCE(f.product,y.product) product,
               COALESCE(f.otdel,y.otdel) otdel,
               COALESCE(f.product_type,y.product_type) product_type,
               COALESCE(f.agent,y.agent) agent, COALESCE(f.orderer,y.orderer) orderer,
               COALESCE(f.courier,y.courier) courier, COALESCE(f.zone,y.zone) zone,
               COALESCE(f.shop_type,y.shop_type) shop_type,
               COALESCE(f.shop_name,y.shop_name) shop_name,
               COALESCE(f.shop_no,y.shop_no) shop_no,
               COALESCE(f.pay_type,y.pay_type) pay_type,
               COALESCE(f.discount_pct,y.discount_pct) discount_pct,
               COALESCE(f.qty,0) fakt_qty, COALESCE(y.qty,0) yak_qty,
               COALESCE(f.amount,0) fakt_summa, COALESCE(y.amount,0) yak_summa
        FROM f FULL JOIN y ON f.sale_date = y.sale_date
          AND f.order_no IS NOT DISTINCT FROM y.order_no
          AND f.product = y.product AND f.otdel = y.otdel
        ORDER BY 1,2
    """, params)

    def metric():
        return {"fakt_summa": 0.0, "yak_summa": 0.0, "kesilgan_summa": 0.0,
                "qoshilgan_summa": 0.0, "fakt_qty": 0.0, "yak_qty": 0.0,
                "kesilgan_qty": 0.0, "qoshilgan_qty": 0.0}

    def add(dst, row):
        fs, ys = float(row["fakt_summa"]), float(row["yak_summa"])
        fq, yq = float(row["fakt_qty"]), float(row["yak_qty"])
        dst["fakt_summa"] += fs; dst["yak_summa"] += ys
        dst["kesilgan_summa"] += max(fs - ys, 0)
        dst["qoshilgan_summa"] += max(ys - fs, 0)
        dst["fakt_qty"] += fq; dst["yak_qty"] += yq
        dst["kesilgan_qty"] += max(fq - yq, 0)
        dst["qoshilgan_qty"] += max(yq - fq, 0)

    if order_no is not None:
        total = metric()
        for row in rows:
            add(total, row)
        info = ({k: rows[0][k] for k in
                 ("order_no", "otdel", "agent", "orderer", "courier", "zone",
                  "shop_type", "shop_name", "shop_no", "pay_type", "discount_pct")}
                if rows else {"order_no": order_no, "otdel": otdel})
        items = []
        for row in rows:
            m = metric(); add(m, row)
            items.append({"product": row["product"], "product_type": row["product_type"],
                          **{k: round(v, 2) for k, v in m.items()}})
        return {"level": "buyurtma", "sana": sana, "info": info,
                "summary": {k: round(v, 2) for k, v in total.items()}, "items": items}

    key = "order_no" if otdel else "otdel"
    grouped = {}
    meta = {}
    for row in rows:
        value = row[key]
        if value not in grouped:
            grouped[value] = metric()
            meta[value] = {"shop_name": row["shop_name"], "agent": row["agent"]}
        add(grouped[value], row)
    items = [{key: value, **meta[value], **{k: round(v, 2) for k, v in values.items()}}
             for value, values in grouped.items()]
    items.sort(key=lambda x: x["kesilgan_summa"], reverse=True)
    return {"level": "buyurtmalar" if otdel else "otdellar", "sana": sana,
            "otdel": otdel, "items": items}


@router.get("/fayllar")
async def fayllar(manba: str = Query("fakt", pattern="^(fakt|yakuniy)$")):
    """Yuklangan ma'lumot: fakt — kun bo'yicha, yakuniy — fayl bo'yicha."""
    if manba == "fakt":
        rows = await db.q("SELECT * FROM v_kunlar ORDER BY sale_date DESC")
        have = {r["sale_date"] for r in rows}
        gaps = []
        if rows:
            d, b = min(have), max(have)
            while d <= b:
                if d.isoweekday() != 7 and d not in have:
                    gaps.append(d.isoformat())
                d += timedelta(days=1)
        return {
            "manba": "fakt", "gaps": gaps, "count": len(rows),
            "items": [{
                "key": r["sale_date"].isoformat(),
                "sana": r["sale_date"].isoformat(),
                "fayl": r["source_file"], "rows": r["qatorlar"],
                "qty": int(r["qty"]), "amount": float(r["amount"]),
                "buyurtma": r["buyurtma"], "dokon": r["dokon"],
                "yuklangan": r["loaded_at"].isoformat(),
            } for r in rows],
        }

    rows = await db.q("""
        SELECT source_file, week_range, min(sale_date) dan, max(sale_date) gacha,
               count(DISTINCT sale_date) kunlar, count(*) qatorlar,
               sum(qty) qty, sum(amount) amount, min(loaded_at) loaded_at
        FROM yakuniy_savdo GROUP BY 1, 2 ORDER BY 3 DESC
    """)
    return {
        "manba": "yakuniy", "gaps": [], "count": len(rows),
        "items": [{
            "key": r["source_file"], "fayl": r["source_file"],
            "sana": f"{r['dan']} … {r['gacha']}",
            "dan": r["dan"].isoformat(), "gacha": r["gacha"].isoformat(),
            "kunlar": r["kunlar"], "rows": r["qatorlar"],
            "qty": int(r["qty"]), "amount": float(r["amount"]),
            "yuklangan": r["loaded_at"].isoformat(),
        } for r in rows],
    }


@router.post("/yukla", dependencies=ADMIN)
async def yukla(
    files: list[UploadFile] = File(...),
    manba: str = Query("fakt", pattern="^(fakt|yakuniy)$"),
    replace: bool = Query(False),
):
    """Excel yuklash. PROGNOZ QAYTA HISOBLANMAYDI."""
    if not files:
        raise HTTPException(400, "Fayl yuborilmadi")

    natija, ok = [], 0
    for f in files:
        name = os.path.basename(f.filename or "nomsiz")
        if not name.lower().endswith((".xlsx", ".xlsm")):
            natija.append({"fayl": name, "status": "error", "manba": None,
                           "xabar": "Faqat .xlsx fayl qabul qilinadi"})
            continue
        try:
            data = await f.read()
            # Excel o'qish bloklovchi — threadpool'da
            rows = await asyncio.to_thread(read_workbook, data)
            kunlar = sorted({r["date"] for r in rows})

            if manba == "fakt" and len(kunlar) > 1:
                raise BadFile(
                    f"Faylda {len(kunlar)} xil sana bor ({kunlar[0]} … {kunlar[-1]}). "
                    "Fakt savdo bitta kunga tegishli bo'lishi kerak. "
                    "Bu yakuniy savdo fayli bo'lsa — manbani «Yakuniy savdo» qiling."
                )

            res = (await load_fakt(name, rows, replace) if manba == "fakt"
                   else await load_yakuniy(name, rows, replace))
        except BadFile as e:
            natija.append({"fayl": name, "status": "error", "manba": None,
                           "xabar": str(e)})
            continue
        except Exception as e:                                   # noqa: BLE001
            natija.append({"fayl": name, "status": "error", "manba": None,
                           "xabar": f"Kutilmagan xato: {e}"})
            continue

        if res["status"] == "skipped":
            natija.append({"fayl": name, "status": "skipped", "manba": manba,
                           "sana": kunlar[0].isoformat(), "xabar": res["reason"]})
            continue

        ok += 1
        row = {"fayl": name, "status": res["status"], "manba": manba,
               "rows": res["rows"], "qty": res["qty"]}
        if manba == "fakt":
            row["sana"] = res["sale_date"].isoformat()
        else:
            row["sana"] = f"{res['dan']} … {res['gacha']}"
            row["kunlar"] = res["kunlar"]
        natija.append(row)

    if ok:
        await db.refresh_views()      # faqat agregatlar; PROGNOZ TEGILMAYDI

    return {
        "natija": natija,
        "jami": {
            "yuklandi": sum(1 for r in natija if r["status"] == "loaded"),
            "qayta": sum(1 for r in natija if r["status"] == "replaced"),
            "otkazildi": sum(1 for r in natija if r["status"] == "skipped"),
            "xato": sum(1 for r in natija if r["status"] == "error"),
        },
        "qayta_hisoblash_kerak": ok > 0,
        "xabar": ("Ma'lumot yuklandi. Prognoz O'ZGARMADI — yangilash uchun "
                  "«Qayta hisoblash» tugmasini bosing.") if ok else None,
    }


@router.delete("/fayllar", dependencies=ADMIN)
async def ochir_hammasi(
    manba: str = Query(..., pattern="^(fakt|yakuniy)$"),
    tasdiq: str = Query(..., description="'HAMMASINI OCHIRISH' deb yozilishi shart"),
):
    """Bitta manbaning BARCHA ma'lumotini o'chiradi. Arxivga tegilmaydi."""
    if tasdiq != "HAMMASINI OCHIRISH":
        raise HTTPException(400, "Tasdiqlash matni noto'g'ri")

    tbl = "fakt_savdo" if manba == "fakt" else "yakuniy_savdo"
    r = await db.q(
        f"SELECT count(*) n, count(DISTINCT sale_date) d, sum(qty) q FROM {tbl}",
        one=True)
    if not r["n"]:
        raise HTTPException(404, f"{manba} savdo bazada yo'q")

    await db.x(f"TRUNCATE {tbl} RESTART IDENTITY")
    await db.refresh_views()
    return {
        "ochirildi": {"manba": manba, "rows": r["n"], "kunlar": r["d"],
                      "qty": int(r["q"] or 0)},
        "xabar": "Prognoz arxiviga tegilmadi. Yangi reja uchun qayta hisoblang.",
    }


@router.delete("/fayllar/{manba}/{key:path}", dependencies=ADMIN)
async def ochir_bitta(manba: str, key: str):
    """fakt — key = sana (2026-07-11); yakuniy — key = fayl nomi."""
    if manba not in ("fakt", "yakuniy"):
        raise HTTPException(400, "manba: fakt yoki yakuniy")

    if manba == "fakt":
        r = await db.q("SELECT count(*) n, sum(qty) q FROM fakt_savdo WHERE sale_date = %s",
                       (key,), one=True)
        if not r["n"]:
            raise HTTPException(404, f"{key} sanasi bazada yo'q")
        await db.x("DELETE FROM fakt_savdo WHERE sale_date = %s", (key,))
        info = {"manba": "fakt", "sana": key, "rows": r["n"], "qty": int(r["q"] or 0)}
    else:
        r = await db.q("""SELECT count(*) n, sum(qty) q, count(DISTINCT sale_date) d,
                                 min(sale_date) a, max(sale_date) b
                          FROM yakuniy_savdo WHERE source_file = %s""",
                       (key,), one=True)
        if not r["n"]:
            raise HTTPException(404, f"«{key}» fayli bazada yo'q")
        await db.x("DELETE FROM yakuniy_savdo WHERE source_file = %s", (key,))
        info = {"manba": "yakuniy", "fayl": key, "rows": r["n"],
                "qty": int(r["q"] or 0), "kunlar": r["d"],
                "dan": r["a"].isoformat(), "gacha": r["b"].isoformat()}

    await db.refresh_views()
    return {"ochirildi": info, "xabar": "Prognoz o'zgarmadi — qayta hisoblash qo'lda."}


@router.get("/chiqarilgan")
async def chiqarilgan():
    """Rejaga kirmagan mahsulotlar va sababi."""
    rows = await db.q("""
        SELECT h.product, h.product_type, h.holat, h.kun_soni,
               h.talab_jami, h.talab_30k, h.oxirgi_sotuv, p.product_id
        FROM v_mahsulot_holati h
        LEFT JOIN products p ON p.name = h.product
        WHERE h.holat <> 'tirik'
        ORDER BY h.talab_jami DESC
    """)
    out = [{
        "product_id": r["product_id"], "name": r["product"], "type": r["product_type"],
        "holat": r["holat"], "kunlar": r["kun_soni"],
        "jami": int(r["talab_jami"] or 0), "yaqinda": int(r["talab_30k"] or 0),
        "oxirgi_sotuv": r["oxirgi_sotuv"].isoformat() if r["oxirgi_sotuv"] else None,
        "sabab": ("Buyurtma bor, lekin ishlab chiqarilmayapti — oxirgi 30 kunda "
                  "yakuniy savdo nol." if r["holat"] == "olik"
                  else f"Siyrak sotuv — atigi {r['kun_soni']} kun tarix "
                       f"(kamida 30 kun kerak)"),
    } for r in rows]

    SABAB = {
        "Тара": "Qadoq (yashik/bidon) — ishlab chiqarish mahsuloti emas",
        "Сырьё": "Xomashyo — reja mahsuloti emas",
    }
    turlar = await db.q("""
        SELECT f.product, f.product_type, count(DISTINCT f.sale_date) kunlar,
               sum(f.qty)::bigint jami,
               sum(f.qty) FILTER (
                   WHERE f.sale_date > (SELECT max(sale_date) - 20 FROM fakt_savdo)
               )::bigint yaqinda, p.product_id
        FROM fakt_savdo f
        LEFT JOIN products p ON p.name = f.product
        WHERE f.product_type IN ('Тара', 'Сырьё')
        GROUP BY 1, 2, p.product_id ORDER BY 4 DESC
    """)
    out += [{
        "product_id": r["product_id"], "name": r["product"], "type": r["product_type"],
        "holat": "tur", "kunlar": r["kunlar"],
        "jami": int(r["jami"] or 0), "yaqinda": int(r["yaqinda"] or 0),
        "oxirgi_sotuv": None, "sabab": SABAB[r["product_type"]],
    } for r in turlar]

    out.sort(key=lambda r: -r["yaqinda"])
    return out


# ═══════════════════════════════════════════════════════════════════════════
#   EXCEL EKSPORT — потребность formatida
# ═══════════════════════════════════════════════════════════════════════════

def _excel(rows, run) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dates = sorted({r["target_date"] for r in rows})
    prods = sorted({r["product"] for r in rows})
    cell = {(r["product"], r["target_date"]): float(r["qty"]) for r in rows}

    all_dates, d = [], dates[0]
    while d <= dates[-1]:
        all_dates.append(d)
        d += timedelta(days=1)

    DN = {1: "1 - понедельник", 2: "2 - вторник", 3: "3 - среда", 4: "4 - четверг",
          5: "5 - пятница", 6: "6 - суббота", 7: "7 - воскресенье"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "подребность"

    hdr = Font(bold=True, size=10)
    thin = Side(style="thin", color="BBBBBB")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="EFF3F5")
    ctr = Alignment(horizontal="center", vertical="center")

    for i, dt in enumerate(all_dates):
        c = ws.cell(row=1, column=2 + i, value=DN[dt.isoweekday()])
        c.font, c.fill, c.border, c.alignment = hdr, fill, bd, ctr
        c = ws.cell(row=2, column=2 + i, value=dt)
        c.number_format = "DD.MM.YYYY"
        c.font, c.fill, c.border, c.alignment = hdr, fill, bd, ctr
    c = ws.cell(row=2, column=1, value="Продукт")
    c.font, c.fill, c.border = hdr, fill, bd

    for j, p in enumerate(prods):
        r = 3 + j
        c = ws.cell(row=r, column=1, value=p)
        c.border = bd
        for i, dt in enumerate(all_dates):
            v = cell.get((p, dt), 0)
            c = ws.cell(row=r, column=2 + i, value=int(round(v / 50) * 50))
            c.border, c.alignment = bd, ctr

    tr = 3 + len(prods)
    c = ws.cell(row=tr, column=1, value="ЖАМИ")
    c.font, c.fill, c.border = hdr, fill, bd
    for i in range(len(all_dates)):
        col = get_column_letter(2 + i)
        c = ws.cell(row=tr, column=2 + i, value=f"=SUM({col}3:{col}{tr - 1})")
        c.font, c.fill, c.border, c.alignment = hdr, fill, bd, ctr

    ws.column_dimensions["A"].width = 44
    for i in range(len(all_dates)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 11
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/eksport")
async def eksport(dokon: str = Query(None),
                  usul: str = Query("aralash", pattern="^(aralash|taqsimot|alohida)$")):
    """Rejani `потребность` formatida Excel'ga chiqaradi."""
    run = await _run()

    if dokon:
        rows = await db.q("""
            SELECT product, target_date, qty FROM fn_reja_dokon(%s)
            WHERE shop_type = %s ORDER BY product, target_date
        """, (usul, dokon))
    else:
        rows = await db.q("""
            SELECT product, target_date, qty FROM v_joriy_reja
            ORDER BY product, target_date
        """)
    if not rows:
        raise HTTPException(404, "Reja bo'sh")

    data = await asyncio.to_thread(_excel, rows, run)
    nom = f"potrebnost_{run['dan']}_{run['gacha']}_run{run['run_id']}"
    if dokon:
        nom += "_" + dokon.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nom}.xlsx"'},
    )
