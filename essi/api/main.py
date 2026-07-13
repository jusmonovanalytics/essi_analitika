"""ESSI — savdo prognozi API.

Model butunlay PostgreSQL da (essi bazasi):
    fn_prognoz()      — bashorat: daraja x mavsum x hafta-kuni x ustama x kalibrovka
    fn_reja_saqla()   — QO'LDA chaqiriladigan yagona hisoblash nuqtasi
    v_joriy_reja      — joriy (faol) reja, arxivdan o'qiladi

Asosiy qoidalar:
  * Yangi Excel yuklanishi prognozni O'ZGARTIRMAYDI. Sayt faqat "eskirgan"
    deb ogohlantiradi; qayta hisoblashni foydalanuvchi bosadi.
  * Har bir hisob arxivga yoziladi. Arxiv o'zgarmas — o'chirib ham,
    tahrirlab ham bo'lmaydi (baza triggeri himoya qiladi).
  * Reja har safar qayta hisoblanmaydi — saqlangan arxivdan o'qiladi.
"""
import io
import json
import os
import sys
from datetime import timedelta

from fastapi import Body, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WEB = os.path.join(ROOT, "web")
sys.path.insert(0, os.path.join(ROOT, "db"))
sys.path.insert(0, os.path.join(ROOT, "api"))

import db                                                          # noqa: E402
from loader import (                                               # noqa: E402
    BadFile, DATA_DIR, YAK_DIR, load_fakt, load_yakuniy, read_workbook,
)

app = FastAPI(title="ESSI Prognoz")

MODEL = "daraja(trim24) x mavsum x hafta-kuni x ustama"


def _run():
    r = db.joriy_run()
    if not r:
        raise HTTPException(
            404,
            "Prognoz hali hisoblanmagan. «Savdo prognozi» bo'limida "
            "«Qayta hisoblash» tugmasini bosing.",
        )
    return r


# ===========================================================================
#  BOSH SAHIFA
# ===========================================================================

@app.get("/api/summary")
def summary():
    run = _run()
    fact = db.q("""
        SELECT count(*) n, sum(qty) qty, sum(amount) amount,
               min(sale_date) d0, max(sale_date) d1, count(DISTINCT sale_date) days
        FROM fakt_savdo
    """, one=True)

    fc = db.q("""
        SELECT sum(qty) qty, min(target_date) d0, max(target_date) d1,
               count(DISTINCT target_date) days, count(DISTINCT product_id) products
        FROM v_joriy_reja
    """, one=True)

    # taqqoslash bazasi: gorizontga teng uzunlikdagi oxirgi haqiqiy davr
    prev = db.q("""
        SELECT sum(qty) qty FROM mv_talab
        WHERE sale_date IN (SELECT DISTINCT sale_date FROM mv_talab
                            ORDER BY 1 DESC LIMIT %s)
    """, (run["gorizont"],), one=True)

    kesim = db.q("""
        SELECT sum(talab) talab, sum(sotilgan) sotilgan, sum(kesim) kesim
        FROM v_kesim
    """, one=True)

    return {
        "run": {
            "run_id": run["run_id"],
            "created_at": run["created_at"].isoformat(),
            "data_last_day": run["data_last_day"].isoformat(),
            "method": MODEL,
            "adjust_factor": float(run["kalibr"]),
            "ustama": run["ustama"],
            "horizon_days": run["gorizont"],
        },
        "fact": {
            "rows": fact["n"], "qty": int(fact["qty"] or 0),
            "amount": float(fact["amount"] or 0),
            "from": fact["d0"].isoformat(), "to": fact["d1"].isoformat(),
            "days": fact["days"],
        },
        "forecast": {
            "qty": float(fc["qty"]), "from": fc["d0"].isoformat(),
            "to": fc["d1"].isoformat(), "days": fc["days"], "products": fc["products"],
        },
        "prev_qty": int(prev["qty"] or 0),
        "wape_14d": float(run["wape"]) if run.get("wape") else None,
        "chegara": 13.69,          # orakul chegarasi — bundan pastga tushib bo'lmaydi
        "kesim": {
            "talab": int(kesim["talab"] or 0),
            "sotilgan": int(kesim["sotilgan"] or 0),
            "yoqotilgan": int(kesim["kesim"] or 0),
            "pct": round(100 * float(kesim["kesim"] or 0) / max(float(kesim["talab"] or 1), 1), 1),
        },
        "stale": db.eskirganmi(),
    }


# ===========================================================================
#  REJA
# ===========================================================================

@app.get("/api/plan")
def plan(ptype: str = Query(None)):
    run = _run()
    horizon = run["gorizont"]

    prev_days = db.q("""
        SELECT DISTINCT sale_date FROM mv_talab ORDER BY 1 DESC LIMIT %s
    """, (horizon,))
    days = [r["sale_date"] for r in prev_days]

    where = "" if not ptype else "AND j.product_type = %(ptype)s"
    rows = db.q(f"""
        WITH prev AS (
            SELECT product, sum(qty) qty FROM mv_talab
            WHERE sale_date = ANY(%(days)s) GROUP BY 1
        )
        SELECT j.product_id, j.product AS name, j.product_type AS type,
               sum(j.qty)                                        AS total,
               sum(j.qty) FILTER (WHERE j.step <= %(half)s)      AS wk1,
               sum(j.qty) FILTER (WHERE j.step >  %(half)s)      AS wk2,
               sum(j.qty_past)                                   AS lo,
               sum(j.qty_yuqori)                                 AS hi,
               max(a.xato)                                       AS xato,
               max(h.kesim_pct)                                  AS kesim,
               max(u.ustama)                                     AS ustama,
               bool_or(m.product IS NOT NULL)                    AS mavsumiy,
               max(p.qty)                                        AS prev_qty
        FROM v_joriy_reja j
        LEFT JOIN v_mahsulot_aniqlik a ON a.product = j.product
        LEFT JOIN v_mahsulot_holati  h ON h.product = j.product
        LEFT JOIN v_ustama           u ON u.product = j.product
        LEFT JOIN (SELECT DISTINCT product FROM mv_mavsum) m ON m.product = j.product
        LEFT JOIN prev p ON p.product = j.product
        WHERE TRUE {where}
        GROUP BY j.product_id, j.product, j.product_type
        ORDER BY total DESC
    """, {"days": days, "half": horizon // 2, "ptype": ptype})

    items = []
    for r in rows:
        prev_q = float(r["prev_qty"]) if r["prev_qty"] else None
        items.append({
            "product_id": r["product_id"], "name": r["name"], "type": r["type"],
            "total": round(float(r["total"])),
            "wk1": round(float(r["wk1"] or 0)), "wk2": round(float(r["wk2"] or 0)),
            "lo": round(float(r["lo"])), "hi": round(float(r["hi"])),
            "wape": round(100 * float(r["xato"]), 1) if r["xato"] else None,
            "kesim": float(r["kesim"]) if r["kesim"] is not None else None,
            "ustama": float(r["ustama"]) if r["ustama"] else 1.0,
            "mavsumiy": bool(r["mavsumiy"]),
            "prev": round(prev_q) if prev_q else None,
            "delta_pct": round(100 * (float(r["total"]) / prev_q - 1), 1) if prev_q else None,
        })

    types = [r["product_type"] for r in db.q("""
        SELECT DISTINCT product_type FROM v_joriy_reja ORDER BY 1
    """)]
    d0 = db.q("SELECT min(target_date) d FROM v_joriy_reja", one=True)["d"]

    return {
        "run_id": run["run_id"], "start": d0.isoformat(), "items": items,
        "types": types, "horizon": horizon,
        "prev_from": min(days).isoformat(), "prev_to": max(days).isoformat(),
    }


@app.get("/api/plan/daily")
def plan_daily(ptype: str = Query(None)):
    _run()
    w = "" if not ptype else "AND product_type = %(ptype)s"
    fc = db.q(f"""
        SELECT target_date d, dow, sum(qty) qty, sum(qty_past) lo, sum(qty_yuqori) hi
        FROM v_joriy_reja WHERE TRUE {w}
        GROUP BY 1, 2 ORDER BY 1
    """, {"ptype": ptype})

    wt = "" if not ptype else "AND product_type = %(ptype)s"
    hist = db.q(f"""
        SELECT sale_date d, dow, sum(qty) qty FROM mv_talab
        WHERE TRUE {wt}
        GROUP BY 1, 2 ORDER BY 1 DESC LIMIT 42
    """, {"ptype": ptype})

    return {
        "history": [{"date": r["d"].isoformat(), "dow": r["dow"], "qty": int(r["qty"])}
                    for r in reversed(hist)],
        "forecast": [{"date": r["d"].isoformat(), "dow": r["dow"],
                      "qty": round(float(r["qty"])), "lo": round(float(r["lo"])),
                      "hi": round(float(r["hi"]))} for r in fc],
    }


USULLAR = {
    "aralash":  ("Aralash (50/50)",  25.13, "Eng aniq. Taqsimot va alohida prognozning o'rtachasi."),
    "taqsimot": ("Taqsimot",         25.51, "So'nggi 24 ish kunidagi haqiqiy ulush bo'yicha bo'lish."),
    "alohida":  ("Alohida prognoz",  25.39, "Har mahsulot × do'kon turi uchun mustaqil model."),
}


@app.get("/api/dokon")
def dokon(usul: str = Query("aralash", regex="^(aralash|taqsimot|alohida)$")):
    """Do'kon turi bo'yicha reja.

    JAMI reja har doim mahsulot darajasidagi modeldan (WAPE 13.75%) — do'kon
    turi faqat o'sha jamini BO'LADI, ya'ni jami hech qachon buzilmaydi.

    Ogohlantirish: do'kon turi darajasida aniqlik ~25% — mahsulot darajasidan
    deyarli ikki barobar yomon. Ishlab chiqarish qarorini JAMI hajmga qarab
    qabul qiling.
    """
    rows = db.q("""
        SELECT shop_type, sum(qty) qty, count(DISTINCT product) mahsulot
        FROM fn_reja_dokon(%s) GROUP BY 1 ORDER BY 2 DESC
    """, (usul,))
    jami = sum(float(r["qty"]) for r in rows) or 1
    nom, wape, izoh = USULLAR[usul]
    return {
        "usul": usul, "usul_nomi": nom, "wape": wape, "izoh": izoh,
        "usullar": [{"kod": k, "nomi": v[0], "wape": v[1], "izoh": v[2]}
                    for k, v in USULLAR.items()],
        "jami": round(jami),
        "mahsulot_wape": 13.75,
        "items": [{
            "shop_type": r["shop_type"], "qty": round(float(r["qty"])),
            "ulush": round(100 * float(r["qty"]) / jami, 1),
            "mahsulot": r["mahsulot"],
        } for r in rows],
    }


@app.get("/api/pivot")
def pivot(ptype: str = Query(None), dokon: str = Query(None),
          usul: str = Query("aralash", regex="^(aralash|taqsimot|alohida)$"),
          round_to: int = Query(50, ge=1, le=500)):
    run = _run()
    w = "" if not ptype else "AND product_type = %(ptype)s"

    if dokon:
        # Do'kon turi kesimi — bu JAMI rejaning ulushi, tahrirlab bo'lmaydi
        rows = db.q(f"""
            SELECT product_id, product, product_type, target_date, qty,
                   qty AS qty_model, FALSE AS ozgartirilgan
            FROM fn_reja_dokon(%(usul)s)
            WHERE shop_type = %(dokon)s {w}
            ORDER BY product_type, product, target_date
        """, {"ptype": ptype, "dokon": dokon, "usul": usul})
    else:
        rows = db.q(f"""
            SELECT product_id, product, product_type, target_date, qty, qty_model,
                   ozgartirilgan
            FROM v_joriy_reja WHERE TRUE {w}
            ORDER BY product_type, product, target_date
        """, {"ptype": ptype})

    cells, model, edited, meta, dates = {}, {}, {}, {}, set()
    for r in rows:
        pid, td = r["product_id"], r["target_date"]
        cells.setdefault(pid, {})[td] = float(r["qty"])
        model.setdefault(pid, {})[td] = float(r["qty_model"] or r["qty"])
        edited.setdefault(pid, {})[td] = bool(r["ozgartirilgan"])
        meta[pid] = (r["product"], r["product_type"])
        dates.add(td)

    if not dates:
        return {"columns": [], "rows": [], "totals": [], "grand_total": 0}

    def rnd(v):
        return int(round(v / round_to) * round_to)

    d0, d1 = min(dates), max(dates)
    col_dates, d = [], d0
    while d <= d1:
        col_dates.append(d)
        d += timedelta(days=1)
    columns = [{"date": c.isoformat(), "dow": c.isoweekday(), "is_forecast": c in dates}
               for c in col_dates]

    out = []
    for pid, (name, pt) in meta.items():
        vals = [rnd(cells[pid][c]) if c in cells[pid] else None for c in col_dates]
        mods = [rnd(model[pid][c]) if c in model[pid] else None for c in col_dates]
        eds = [bool(edited[pid].get(c)) if c in cells[pid] else False for c in col_dates]
        out.append({"product_id": pid, "name": name, "type": pt, "values": vals,
                    "model": mods, "edited": eds,
                    "total": sum(v for v in vals if v)})
    out.sort(key=lambda r: r["name"])

    totals = [sum(r["values"][i] or 0 for r in out) if c["is_forecast"] else None
              for i, c in enumerate(columns)]
    return {"run_id": run["run_id"], "round_to": round_to, "columns": columns,
            "rows": out, "totals": totals, "grand_total": sum(r["total"] for r in out),
            "qolda": run.get("qolda", False), "asos_run": run.get("asos_run"),
            "dokon": dokon, "usul": usul if dokon else None,
            # do'kon turi kesimida tahrirlash MUMKIN EMAS — hujayrada jami
            # rejaning ulushi turadi, uni o'zgartirish ma'nosiz
            "tahrirlanadi": dokon is None}


@app.get("/api/product/{pid}")
def product(pid: int):
    p = db.q("SELECT product_id, name, product_type FROM products WHERE product_id = %s",
             (pid,), one=True)
    if not p:
        raise HTTPException(404, "Mahsulot topilmadi")

    hist = db.q("""
        SELECT t.sale_date d, t.dow, sum(t.qty) qty, sum(t.amount) amount,
               max(s.qty) sotilgan
        FROM mv_talab t
        LEFT JOIN mv_sotilgan s ON s.product = t.product AND s.sale_date = t.sale_date
        WHERE t.product = %s GROUP BY 1, 2 ORDER BY 1
    """, (p["name"],))

    fc = db.q("""
        SELECT target_date d, dow, step, qty, qty_past lo, qty_yuqori hi,
               daraja, mavsum, dow_ix
        FROM v_joriy_reja WHERE product_id = %s ORDER BY target_date
    """, (pid,))

    dow = db.q("""
        SELECT dow, round(avg(qty)) q FROM (
            SELECT dow, qty FROM mv_talab_zich
            WHERE product = %s ORDER BY sale_date DESC LIMIT 48
        ) t GROUP BY 1 ORDER BY 1
    """, (p["name"],))

    seas = db.q("SELECT oy, ix FROM mv_mavsum WHERE product = %s ORDER BY oy", (p["name"],))
    acc = db.q("SELECT * FROM v_mahsulot_aniqlik WHERE product = %s", (p["name"],), one=True)
    st = db.q("SELECT * FROM v_mahsulot_holati WHERE product = %s", (p["name"],), one=True)

    return {
        "product": dict(p),
        "history": [{"date": r["d"].isoformat(), "dow": r["dow"], "qty": int(r["qty"]),
                     "amount": float(r["amount"]),
                     "sotilgan": int(r["sotilgan"]) if r["sotilgan"] is not None else None}
                    for r in hist],
        "forecast": [{"date": r["d"].isoformat(), "dow": r["dow"], "step": r["step"],
                      "qty": round(float(r["qty"]), 1), "lo": round(float(r["lo"]), 1),
                      "hi": round(float(r["hi"]), 1), "daraja": float(r["daraja"] or 0),
                      "mavsum": float(r["mavsum"] or 1), "dow_ix": float(r["dow_ix"] or 1)}
                     for r in fc],
        "dow_profile": [{"dow": r["dow"], "qty": round(float(r["q"]))} for r in dow],
        "seasonality": [{"oy": r["oy"], "ix": float(r["ix"])} for r in seas],
        "accuracy": {
            "cv": float(acc["cv"]) if acc and acc["cv"] else None,
            "xato_pct": round(100 * float(acc["xato"]), 1) if acc and acc["xato"] else None,
            "n_kun": acc["n_kun"] if acc else None,
        } if acc else None,
        "holat": {
            "holat": st["holat"], "kesim_pct": float(st["kesim_pct"] or 0),
            "talab_jami": float(st["talab_jami"] or 0),
            "sotilgan_jami": float(st["sotilgan_jami"] or 0),
        } if st else None,
    }


# ===========================================================================
#  KESIM — yo'qotilgan savdo
# ===========================================================================

@app.get("/api/kesim")
def kesim(limit: int = Query(50, ge=1, le=200)):
    """Ombor yetmagani uchun kesilgan savdo — mahsulot kesimida."""
    rows = db.q("""
        SELECT h.product, h.product_type, h.talab_jami, h.sotilgan_jami,
               h.talab_jami - h.sotilgan_jami AS kesim,
               h.kesim_pct, h.holat, u.ustama,
               p.product_id
        FROM v_mahsulot_holati h
        JOIN v_ustama u USING (product)
        LEFT JOIN products p ON p.name = h.product
        WHERE h.talab_jami > 0
        ORDER BY (h.talab_jami - h.sotilgan_jami) DESC
        LIMIT %s
    """, (limit,))

    jami = db.q("""
        SELECT sum(talab) t, sum(sotilgan) s, sum(kesim) k,
               count(DISTINCT sale_date) kunlar
        FROM v_kesim
    """, one=True)

    # narx orqali pul bahosi
    narx = db.q("""
        SELECT sum(amount) / NULLIF(sum(qty), 0) AS narx FROM fakt_savdo
    """, one=True)
    n = float(narx["narx"] or 0)

    return {
        "jami": {
            "talab": int(jami["t"] or 0),
            "sotilgan": int(jami["s"] or 0),
            "kesim": int(jami["k"] or 0),
            "pct": round(100 * float(jami["k"] or 0) / max(float(jami["t"] or 1), 1), 1),
            "kunlar": jami["kunlar"],
            "pul": round(float(jami["k"] or 0) * n),
            "ort_narx": round(n),
        },
        "items": [{
            "product_id": r["product_id"], "name": r["product"], "type": r["product_type"],
            "talab": int(r["talab_jami"]), "sotilgan": int(r["sotilgan_jami"]),
            "kesim": int(r["kesim"]), "kesim_pct": float(r["kesim_pct"] or 0),
            "holat": r["holat"], "ustama": float(r["ustama"]),
            "pul": round(float(r["kesim"]) * n),
        } for r in rows],
    }


@app.get("/api/kesim/daily")
def kesim_daily():
    """Kesimning kunlik dinamikasi."""
    rows = db.q("""
        SELECT sale_date d, sum(talab) t, sum(sotilgan) s, sum(kesim) k
        FROM v_kesim GROUP BY 1 ORDER BY 1
    """)
    return [{"date": r["d"].isoformat(), "talab": int(r["t"]),
             "sotilgan": int(r["s"]), "kesim": int(r["k"]),
             "pct": round(100 * float(r["k"]) / max(float(r["t"]), 1), 1)} for r in rows]


# ===========================================================================
#  MA'LUMOT
# ===========================================================================

@app.get("/api/dow")
def dow():
    rows = db.q("""
        SELECT dow, count(*) days, round(avg(qty)) qty, round(avg(amount)) amount
        FROM v_kunlik_jami GROUP BY 1 ORDER BY 1
    """)
    if not rows:
        return []
    avg = sum(float(r["qty"]) for r in rows) / len(rows)
    return [{"dow": r["dow"], "days": r["days"], "qty": int(r["qty"]),
             "orders": 0, "index": round(float(r["qty"]) / avg, 3)} for r in rows]


@app.get("/api/excluded")
def excluded():
    """Rejaga kirmagan mahsulotlar va sababi. Ro'yxat qaytaradi (sayt shuni kutadi)."""
    rows = db.q("""
        SELECT h.product, h.product_type, h.holat, h.kun_soni,
               h.talab_jami, h.talab_30k, h.sotilgan_30k, h.oxirgi_sotuv,
               p.product_id
        FROM v_mahsulot_holati h
        LEFT JOIN products p ON p.name = h.product
        WHERE h.holat <> 'tirik'
        ORDER BY h.talab_jami DESC
    """)
    out = []
    for r in rows:
        if r["holat"] == "olik":
            reason = ("Buyurtma bor, lekin ishlab chiqarilmayapti — oxirgi 30 kunda "
                      "yakuniy savdo nol. Rejaga kiritilmaydi.")
        else:
            reason = f"Siyrak sotuv — atigi {r['kun_soni']} kun tarix (kamida 30 kun kerak)"
        out.append({
            "product_id": r["product_id"], "name": r["product"], "type": r["product_type"],
            "holat": r["holat"], "n_days": r["kun_soni"],
            "total_qty": int(r["talab_jami"] or 0),
            "recent_qty": int(r["talab_30k"] or 0),
            "oxirgi_sotuv": r["oxirgi_sotuv"].isoformat() if r["oxirgi_sotuv"] else None,
            "reason": reason,
        })

    # Butun turi bilan chiqarilganlar — ular ham ro'yxatga
    # (Мехрибон 2026-07-13 dan REJAGA KIRADI — bu ro'yxatda emas)
    SABAB = {
        "Тара": "Qadoq (yashik/bidon) — ishlab chiqarish mahsuloti emas",
        "Сырьё": "Xomashyo — reja mahsuloti emas",
    }
    turlar = db.q("""
        SELECT f.product, f.product_type,
               count(DISTINCT f.sale_date)  n_days,
               sum(f.qty)::bigint           total_qty,
               sum(f.qty) FILTER (
                   WHERE f.sale_date > (SELECT max(sale_date) - 20 FROM fakt_savdo)
               )::bigint                    recent_qty,
               p.product_id
        FROM fakt_savdo f
        LEFT JOIN products p ON p.name = f.product
        WHERE f.product_type IN ('Тара', 'Сырьё')
        GROUP BY 1, 2, p.product_id
        ORDER BY 4 DESC
    """)
    for r in turlar:
        out.append({
            "product_id": r["product_id"], "name": r["product"], "type": r["product_type"],
            "holat": "tur", "n_days": r["n_days"],
            "total_qty": int(r["total_qty"] or 0),
            "recent_qty": int(r["recent_qty"] or 0),
            "oxirgi_sotuv": None,
            "reason": SABAB[r["product_type"]],
        })

    out.sort(key=lambda r: -r["recent_qty"])
    return out


@app.get("/api/overview")
def overview():
    f = db.q("""
        SELECT count(*) rows, sum(qty) qty, sum(amount) amount,
               min(sale_date) d0, max(sale_date) d1, count(DISTINCT sale_date) days,
               count(DISTINCT order_no) orders, count(DISTINCT source_file) files
        FROM fakt_savdo
    """, one=True)
    y = db.q("""
        SELECT count(*) rows, sum(qty) qty, min(sale_date) d0, max(sale_date) d1,
               count(DISTINCT sale_date) days, count(DISTINCT source_file) files
        FROM yakuniy_savdo
    """, one=True)
    d = db.q("""
        SELECT (SELECT count(*) FROM products)                       products,
               (SELECT count(DISTINCT shop_no) FROM fakt_savdo)      shops,
               (SELECT count(DISTINCT zone) FROM fakt_savdo)         zones,
               (SELECT count(DISTINCT agent) FROM fakt_savdo)        agents,
               (SELECT count(DISTINCT courier) FROM fakt_savdo)      couriers
    """, one=True)
    size = db.q("SELECT pg_size_pretty(pg_database_size(current_database())) sz", one=True)

    return {
        "sales": {
            "rows": f["rows"], "qty": int(f["qty"] or 0), "amount": float(f["amount"] or 0),
            "orders": f["orders"], "days": f["days"], "files": f["files"],
            "from": f["d0"].isoformat() if f["d0"] else None,
            "to": f["d1"].isoformat() if f["d1"] else None,
        },
        "yakuniy": {
            "rows": y["rows"], "qty": int(y["qty"] or 0), "days": y["days"],
            "files": y["files"],
            "from": y["d0"].isoformat() if y["d0"] else None,
            "to": y["d1"].isoformat() if y["d1"] else None,
        },
        "dims": dict(d), "db_size": size["sz"],
    }


@app.get("/api/batches")
def batches(manba: str = Query("fakt", regex="^(fakt|yakuniy)$")):
    """Yuklangan ma'lumot ro'yxati.

    fakt    — kun bo'yicha (har kun = bitta fayl)
    yakuniy — fayl bo'yicha (har fayl = bir necha kun)
    """
    if manba == "fakt":
        rows = db.q("SELECT * FROM v_kunlar ORDER BY sale_date DESC")
        have = {r["sale_date"] for r in rows}
        gaps = []
        if rows:
            d, b = min(have), max(have)
            while d <= b:
                if d.isoweekday() != 7 and d not in have:
                    gaps.append(d.isoformat())
                d += timedelta(days=1)
        return {
            "manba": "fakt",
            "batches": [{
                "key": r["sale_date"].isoformat(),
                "sale_date": r["sale_date"].isoformat(),
                "source_file": r["source_file"],
                "rows": r["qatorlar"], "qty": int(r["qty"]),
                "amount": float(r["amount"]),
                "orders": r["buyurtma"], "shops": r["dokon"],
                "loaded_at": r["loaded_at"].isoformat(),
            } for r in rows],
            "gaps": gaps, "count": len(rows),
        }

    rows = db.q("""
        SELECT source_file, week_range,
               min(sale_date) dan, max(sale_date) gacha,
               count(DISTINCT sale_date) kunlar,
               count(*) qatorlar, sum(qty) qty, sum(amount) amount,
               min(loaded_at) loaded_at
        FROM yakuniy_savdo GROUP BY 1, 2 ORDER BY 3 DESC
    """)
    return {
        "manba": "yakuniy",
        "batches": [{
            "key": r["source_file"],
            "source_file": r["source_file"], "week_range": r["week_range"],
            "sale_date": f"{r['dan']} … {r['gacha']}",
            "dan": r["dan"].isoformat(), "gacha": r["gacha"].isoformat(),
            "kunlar": r["kunlar"], "rows": r["qatorlar"],
            "qty": int(r["qty"]), "amount": float(r["amount"]),
            "loaded_at": r["loaded_at"].isoformat(),
        } for r in rows],
        "gaps": [], "count": len(rows),
    }


# ===========================================================================
#  ARXIV — prognoz tarixi
# ===========================================================================

@app.get("/api/runs")
def runs(limit: int = Query(50, ge=1, le=200)):
    """Prognoz arxivi. Hech qachon o'chirilmaydi."""
    rows = db.q("SELECT * FROM v_arxiv LIMIT %s", (limit,))
    return [{
        "run_id": r["run_id"], "created_at": r["created_at"].isoformat(),
        "faol": r["faol"], "data_last_day": r["data_last_day"].isoformat(),
        "horizon": r["gorizont"], "ustama": r["ustama"], "adjust": float(r["kalibr"]),
        "products": r["n_mahsulot"], "total_qty": float(r["jami_qty"]),
        "from": r["dan"].isoformat(), "to": r["gacha"].isoformat(),
        "fakt_kunlar": r["fakt_kunlar"], "fakt_fayllar": r["fakt_fayllar"],
        "yak_kunlar": r["yak_kunlar"],
        "qolda": r["qolda"], "asos_run": r["asos_run"],
        "ozgartirilgan": r["ozgartirilgan_qator"],
        "farq": float(r["farq_oldingidan"]) if r["farq_oldingidan"] else None,
        "notes": r["izoh"],
    } for r in rows]


@app.get("/api/runs/{run_id}")
def run_detail(run_id: int):
    """Arxivdagi bitta prognozni to'liq ko'rish."""
    r = db.q("SELECT * FROM v_arxiv WHERE run_id = %s", (run_id,), one=True)
    if not r:
        raise HTTPException(404, f"run_id={run_id} arxivda yo'q")
    items = db.q("""
        SELECT p.product_id, p.name, p.product_type,
               sum(d.qty) total, sum(d.qty_past) lo, sum(d.qty_yuqori) hi
        FROM reja_daily d JOIN products p USING (product_id)
        WHERE d.run_id = %s
        GROUP BY 1, 2, 3 ORDER BY total DESC
    """, (run_id,))
    return {
        "run": {"run_id": r["run_id"], "created_at": r["created_at"].isoformat(),
                "faol": r["faol"], "data_last_day": r["data_last_day"].isoformat(),
                "total_qty": float(r["jami_qty"]), "notes": r["izoh"]},
        "items": [{"product_id": i["product_id"], "name": i["name"],
                   "type": i["product_type"], "total": round(float(i["total"])),
                   "lo": round(float(i["lo"])), "hi": round(float(i["hi"]))}
                  for i in items],
    }


@app.post("/api/plan/edit")
def plan_edit(payload: dict = Body(...)):
    """Rejani QO'LDA tahrirlash.

    Eski reja O'ZGARMAYDI — undan yangi versiya yaratiladi. Modelning asl
    qiymati (`qty_model`) saqlanadi, shunda nimani o'zgartirganingiz doim ko'rinadi.

    body: {
      "changes": [{"product_id": 1, "target_date": "2026-07-13", "qty": 500}, ...],
      "asos": 6,            # ixtiyoriy — qaysi rejadan (NULL = joriy)
      "izoh": "..."         # ixtiyoriy
    }
    """
    changes = payload.get("changes") or []
    if not changes:
        raise HTTPException(400, "O'zgarish yuborilmadi")
    for ch in changes:
        if not all(k in ch for k in ("product_id", "target_date", "qty")):
            raise HTTPException(400, "Har bir o'zgarishda product_id, target_date, qty bo'lishi kerak")
        if float(ch["qty"]) < 0:
            raise HTTPException(400, "Manfiy miqdor bo'lishi mumkin emas")

    try:
        run_id = db.x(
            "SELECT fn_reja_qolda(%s::jsonb, %s, %s)",
            (json.dumps(changes), payload.get("asos"), payload.get("izoh")),
        )
    except Exception as e:                                   # noqa: BLE001
        raise HTTPException(400, str(e).split("CONTEXT")[0].strip())

    r = db.q("SELECT * FROM v_arxiv WHERE run_id = %s", (run_id,), one=True)
    return {
        "ok": True, "run_id": run_id, "asos_run": r["asos_run"],
        "jami_qty": float(r["jami_qty"]),
        "ozgartirilgan": r["ozgartirilgan_qator"],
        "farq": float(r["farq_oldingidan"]) if r["farq_oldingidan"] else 0,
    }


@app.post("/api/runs/{run_id}/activate")
def activate(run_id: int):
    """Arxivdagi eski rejaga qaytish."""
    try:
        msg = db.x("SELECT fn_reja_faollashtir(%s)", (run_id,))
    except Exception as e:
        raise HTTPException(400, str(e))
    return {"ok": True, "message": msg}


@app.post("/api/recompute")
def recompute(
    horizon: int = Query(12, ge=6, le=30),
    ustama: bool = Query(True),
    adjust: float = Query(1.03, ge=0.5, le=1.5),
    izoh: str = Query(None),
):
    """QAYTA HISOBLASH — faqat shu yerda. Avtomatik chaqirilmaydi.

    Eski reja arxivda qoladi, yangisi uning yoniga qo'shiladi.
    """
    eski = db.joriy_run()
    run_id = db.x("SELECT fn_reja_saqla(%s, %s, %s, %s)",
                  (horizon, ustama, adjust, izoh))
    yangi = db.q("SELECT * FROM reja_runs WHERE run_id = %s", (run_id,), one=True)
    return {
        "ok": True,
        "run_id": run_id,
        "jami_qty": float(yangi["jami_qty"]),
        "n_mahsulot": yangi["n_mahsulot"],
        "dan": yangi["dan"].isoformat(),
        "gacha": yangi["gacha"].isoformat(),
        "arxivlangan": {
            "run_id": eski["run_id"],
            "jami_qty": float(eski["jami_qty"]),
        } if eski else None,
        "farq": (float(yangi["jami_qty"]) - float(eski["jami_qty"])) if eski else None,
    }


# ===========================================================================
#  YUKLASH — prognozni O'ZGARTIRMAYDI
# ===========================================================================

@app.post("/api/upload")
async def upload(
    files: list[UploadFile] = File(...),
    manba: str = Query("auto", regex="^(auto|fakt|yakuniy)$"),
    replace: bool = Query(False),
):
    """Excel hisobotlarini yuklaydi — fakt savdo yoki yakuniy savdo.

    Ikkala manba ham bir xil 16 ustunli formatda keladi. Farqi — kunlar soni:
        fakt savdo    — bitta kun (kunlik hisobot)
        yakuniy savdo — bir necha kun (haftalik hisobot)

    `manba=auto` (standart) — fayl ichidagi sanalar soniga qarab o'zi aniqlaydi.
    `manba=fakt` / `manba=yakuniy` — majburiy tanlov.

    DIQQAT: prognoz QAYTA HISOBLANMAYDI. Javobda `needs_recompute` bayrog'i
    qaytadi — foydalanuvchi o'zi qaror qiladi.
    """
    if not files:
        raise HTTPException(400, "Fayl yuborilmadi")

    results, ok = [], 0
    for f in files:
        name = os.path.basename(f.filename or "nomsiz")
        if not name.lower().endswith((".xlsx", ".xlsm")):
            results.append({"file": name, "status": "error", "manba": None,
                            "message": "Faqat .xlsx fayl qabul qilinadi"})
            continue
        try:
            data = await f.read()
            rows = read_workbook(data)

            kunlar = sorted({r["date"] for r in rows})
            tur = manba if manba != "auto" else ("fakt" if len(kunlar) == 1 else "yakuniy")

            if tur == "fakt" and len(kunlar) > 1:
                raise BadFile(
                    f"Faylda {len(kunlar)} xil sana bor ({kunlar[0]} … {kunlar[-1]}). "
                    "Fakt savdo bitta kunga tegishli bo'lishi kerak. "
                    "Bu yakuniy savdo fayli bo'lsa — manbani «Yakuniy savdo» qilib tanlang."
                )

            if tur == "fakt":
                res = load_fakt(name, rows, replace=replace)
            else:
                res = load_yakuniy(name, rows, replace=replace)

        except BadFile as e:
            results.append({"file": name, "status": "error", "manba": None,
                            "message": str(e)})
            continue
        except Exception as e:                       # noqa: BLE001
            results.append({"file": name, "status": "error", "manba": None,
                            "message": f"Kutilmagan xato: {e}"})
            continue

        if res["status"] == "skipped":
            results.append({"file": name, "status": "skipped", "manba": tur,
                            "sale_date": kunlar[0].isoformat(),
                            "message": res["reason"]})
            continue

        ok += 1
        row = {"file": name, "status": res["status"], "manba": tur,
               "rows": res["rows"], "qty": res["qty"]}
        if tur == "fakt":
            row["sale_date"] = res["sale_date"].isoformat()
        else:
            row["sale_date"] = f"{res['dan']} … {res['gacha']}"
            row["kunlar"] = res["kunlar"]
        results.append(row)

        # manba faylini papkaga saqlaymiz — keyingi to'liq qayta yuklash uchun
        try:
            folder = DATA_DIR if tur == "fakt" else YAK_DIR
            os.makedirs(folder, exist_ok=True)
            dst = os.path.join(folder, name)
            if not os.path.exists(dst):
                with open(dst, "wb") as out:
                    out.write(data)
        except OSError:
            pass

    if ok:
        db.refresh_views()          # faqat agregatlar; PROGNOZ TEGILMAYDI

    fk = db.q("SELECT count(*) n, max(sale_date) d FROM fakt_savdo", one=True)
    yk = db.q("SELECT count(*) n, max(sale_date) d FROM yakuniy_savdo", one=True)
    return {
        "results": results,
        "summary": {
            "loaded": sum(1 for r in results if r["status"] == "loaded"),
            "replaced": sum(1 for r in results if r["status"] == "replaced"),
            "skipped": sum(1 for r in results if r["status"] == "skipped"),
            "errors": sum(1 for r in results if r["status"] == "error"),
            "fakt": sum(1 for r in results if r.get("manba") == "fakt"
                        and r["status"] in ("loaded", "replaced")),
            "yakuniy": sum(1 for r in results if r.get("manba") == "yakuniy"
                           and r["status"] in ("loaded", "replaced")),
        },
        "db": {
            "rows": fk["n"], "last_day": fk["d"].isoformat() if fk["d"] else None,
            "yak_rows": yk["n"],
            "yak_last_day": yk["d"].isoformat() if yk["d"] else None,
        },
        "needs_recompute": ok > 0,
        "message": ("Ma'lumot yuklandi. Prognoz O'ZGARMADI — yangilash uchun "
                    "«Qayta hisoblash» tugmasini bosing.") if ok else None,
    }


@app.delete("/api/batches")
def delete_all(
    manba: str = Query(..., regex="^(fakt|yakuniy)$"),
    confirm: str = Query(..., description="'HAMMASINI OCHIRISH' deb yozilishi shart"),
):
    """Bitta manbaning BARCHA ma'lumotini o'chiradi. Ataylab qiyinlashtirilgan.

    Prognoz ARXIVIGA tegilmaydi — u o'zgarmas.
    """
    if confirm != "HAMMASINI OCHIRISH":
        raise HTTPException(400, "Tasdiqlash matni noto'g'ri")

    tbl = "fakt_savdo" if manba == "fakt" else "yakuniy_savdo"
    r = db.q(f"SELECT count(*) n, count(DISTINCT sale_date) d, sum(qty) q FROM {tbl}",
             one=True)
    if not r["n"]:
        raise HTTPException(404, f"{manba} savdo bazada yo'q")

    db.x(f"TRUNCATE {tbl} RESTART IDENTITY")
    db.refresh_views()
    return {
        "deleted": {"manba": manba, "rows": r["n"], "days": r["d"],
                    "qty": int(r["q"] or 0)},
        "message": "Prognoz arxiviga tegilmadi. Yangi reja uchun qayta hisoblang.",
    }


@app.delete("/api/batches/{manba}/{key:path}")
def delete_one(manba: str, key: str):
    """Bitta kunni (fakt) yoki bitta faylni (yakuniy) o'chiradi.

    fakt    — key = sana (2026-07-11)
    yakuniy — key = fayl nomi (08-11-07-2026.xlsx)
    """
    if manba not in ("fakt", "yakuniy"):
        raise HTTPException(400, "manba: fakt yoki yakuniy")

    if manba == "fakt":
        r = db.q("SELECT count(*) n, sum(qty) q FROM fakt_savdo WHERE sale_date = %s",
                 (key,), one=True)
        if not r["n"]:
            raise HTTPException(404, f"{key} sanasi bazada yo'q")
        db.x("DELETE FROM fakt_savdo WHERE sale_date = %s", (key,))
        info = {"manba": "fakt", "sale_date": key,
                "rows": r["n"], "qty": int(r["q"] or 0)}
    else:
        r = db.q("""SELECT count(*) n, sum(qty) q, count(DISTINCT sale_date) d,
                           min(sale_date) a, max(sale_date) b
                    FROM yakuniy_savdo WHERE source_file = %s""", (key,), one=True)
        if not r["n"]:
            raise HTTPException(404, f"«{key}» fayli bazada yo'q")
        db.x("DELETE FROM yakuniy_savdo WHERE source_file = %s", (key,))
        info = {"manba": "yakuniy", "source_file": key,
                "rows": r["n"], "qty": int(r["q"] or 0), "kunlar": r["d"],
                "dan": r["a"].isoformat(), "gacha": r["b"].isoformat()}

    db.refresh_views()
    return {"deleted": info,
            "message": "Prognoz o'zgarmadi — qayta hisoblash qo'lda."}


# ===========================================================================
#  EXCEL EKSPORT
# ===========================================================================

@app.get("/api/export/excel")
def export_excel():
    """Rejani `потребность` formatida Excel'ga chiqaradi."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    run = _run()
    rows = db.q("""
        SELECT product, target_date, qty FROM v_joriy_reja ORDER BY product, target_date
    """)
    dates = sorted({r["target_date"] for r in rows})
    prods = sorted({r["product"] for r in rows})
    cell = {(r["product"], r["target_date"]): float(r["qty"]) for r in rows}

    # yakshanbalarni ham ustun sifatida qo'shamiz (Excel'dagi kabi, 0 bilan)
    all_dates, d = [], dates[0]
    while d <= dates[-1]:
        all_dates.append(d)
        d += timedelta(days=1)

    DN = {1: "1 - понедельник", 2: "2 - вторник", 3: "3 - среда", 4: "4 - четверг",
          5: "5 - пятница", 6: "6 - суббота", 7: "7 - воскресенье"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "подребность"

    hdr = Font(bold=True, size=10)
    thin = Side(style="thin", color="BBBBBB")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="EFF3F5")
    ctr = Alignment(horizontal="center", vertical="center")

    for i, dt in enumerate(all_dates):
        c = ws.cell(row=1, column=2 + i, value=DN[dt.isoweekday()])
        c.font = hdr; c.fill = fill; c.border = bd; c.alignment = ctr
        c = ws.cell(row=2, column=2 + i, value=dt)
        c.number_format = "DD.MM.YYYY"; c.font = hdr; c.fill = fill
        c.border = bd; c.alignment = ctr
    c = ws.cell(row=2, column=1, value="Продукт")
    c.font = hdr; c.fill = fill; c.border = bd

    for j, p in enumerate(prods):
        r = 3 + j
        c = ws.cell(row=r, column=1, value=p)
        c.border = bd
        for i, dt in enumerate(all_dates):
            v = cell.get((p, dt), 0)
            c = ws.cell(row=r, column=2 + i, value=int(round(v / 50) * 50))
            c.border = bd
            c.alignment = ctr

    tr = 3 + len(prods)
    c = ws.cell(row=tr, column=1, value="ЖАМИ")
    c.font = hdr; c.fill = fill; c.border = bd
    for i in range(len(all_dates)):
        col = 2 + i
        letter = openpyxl.utils.get_column_letter(col)
        c = ws.cell(row=tr, column=col,
                    value=f"=SUM({letter}3:{letter}{tr - 1})")
        c.font = hdr; c.fill = fill; c.border = bd; c.alignment = ctr

    ws.column_dimensions["A"].width = 44
    for i in range(len(all_dates)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + i)].width = 11
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"potrebnost_{run['dan']}_{run['gacha']}_run{run['run_id']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ===========================================================================
@app.get("/")
def index():
    return FileResponse(os.path.join(WEB, "index.html"))


app.mount("/static", StaticFiles(directory=WEB), name="static")
