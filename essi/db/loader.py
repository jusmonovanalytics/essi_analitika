"""Excel fakt savdo fayllarini o'qish va `essi` bazasiga yuklash.

Buni ikkalasi ham ishlatadi:
  - db/load_excel.py  (buyruq qatori orqali papkani yuklash)
  - api/main.py       (sayt orqali fayl yuklash)

MUHIM: yuklash prognozni QAYTA HISOBLAMAYDI. Agregatlar yangilanadi, prognoz
esa eski holicha qoladi — toki foydalanuvchi qo'lda qayta hisoblamaguncha.
"""
import io
import os

import openpyxl
import psycopg2

DSN = os.environ.get(
    "ESSI_DSN",
    "host=localhost port=5432 dbname=essi user=postgres password=postgres123 "
    "client_encoding=UTF8",
)
DATA_DIR = os.environ.get("ESSI_DATA_DIR", r"D:\ESSI\essi\fakt savdo")
YAK_DIR = os.environ.get("ESSI_YAK_DIR", r"D:\ESSI\essi\yakuniy savdo")

# Excel ustun indekslari (0-ustun — tartib raqami)
C_DATE, C_AGENT, C_ORDERER, C_COURIER = 1, 2, 3, 4
C_ZONE, C_SHOPTYPE, C_SHOP, C_SHOPNO = 5, 6, 7, 8
C_PTYPE, C_PRODUCT, C_ORDERNO = 9, 10, 11
C_PAY, C_DISC, C_QTY, C_AMOUNT = 12, 13, 14, 15

EXPECTED_HEADER = (
    "По дням", "Ответственный агент магазина", "Заказ оформил", "Доставщик",
    "Зона", "Тип магазина", "Магазин", "№ Магазина", "Тип продукта", "Продукт",
    "№ заказа", "Тип оплаты", "Процент скидки", "Колв.продуктов БВ", "Общ.сумма БВ",
)

FAKT_COLS = ("sale_date", "agent", "orderer", "courier", "zone", "shop_type",
             "shop_name", "shop_no", "product_type", "product", "order_no",
             "pay_type", "discount_pct", "qty", "amount", "source_file")

YAK_COLS = ("sale_date", "week_range", "agent", "orderer", "courier", "zone",
            "shop_type", "shop_name", "shop_no", "product_type", "product",
            "order_no", "pay_type", "discount_pct", "qty", "amount", "source_file")


class BadFile(Exception):
    """Fayl kutilgan formatda emas."""


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def to_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


def to_num(v):
    """Sonli qiymat; 'Не указано' kabi matn bo'lsa — None."""
    if v is None:
        return None
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def esc(v):
    if v is None:
        return r"\N"
    return (str(v).replace("\\", "\\\\").replace("\t", " ")
            .replace("\n", " ").replace("\r", " "))


def read_workbook(src):
    """src: fayl yo'li yoki bytes -> qatorlar ro'yxati.

    Ustunlar tekshiriladi — noto'g'ri fayl bazaga tushmasligi uchun.
    Fakt savdo (bir kun) ham, yakuniy savdo (bir necha kun) ham shu formatda.
    """
    if isinstance(src, (bytes, bytearray)):
        src = io.BytesIO(src)
    try:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:                                       # noqa: BLE001
        raise BadFile(f"Excel fayl sifatida ochib bo'lmadi: {e}")

    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        raise BadFile("Fayl bo'sh")

    got = tuple(norm(h) for h in header[1:16])
    if got != EXPECTED_HEADER:
        missing = [h for h in EXPECTED_HEADER if h not in got]
        raise BadFile(
            "Ustunlar mos kelmadi. Kutilgan: savdo hisoboti (16 ustun). "
            + (f"Topilmadi: {', '.join(missing[:4])}" if missing
               else "Ustunlar tartibi boshqacha.")
        )

    rows = []
    for i, r in enumerate(it, start=2):
        if r[C_DATE] is None or r[C_PRODUCT] is None:
            continue
        try:
            rows.append({
                "date": r[C_DATE].date() if hasattr(r[C_DATE], "date") else r[C_DATE],
                "agent": norm(r[C_AGENT]),
                "orderer": norm(r[C_ORDERER]),
                "courier": norm(r[C_COURIER]),
                "zone": norm(r[C_ZONE]),
                "shop_type": norm(r[C_SHOPTYPE]),
                "shop_name": norm(r[C_SHOP]),
                "shop_no": to_int(r[C_SHOPNO]),
                "product_type": norm(r[C_PTYPE]),
                "product": norm(r[C_PRODUCT]),
                "order_no": to_int(r[C_ORDERNO]),
                "pay_type": norm(r[C_PAY]) or "unknown",
                "discount_pct": to_num(r[C_DISC]),
                "qty": to_num(r[C_QTY]) or 0.0,
                "amount": to_num(r[C_AMOUNT]) or 0.0,
            })
        except (TypeError, ValueError) as e:
            raise BadFile(f"{i}-qatorda noto'g'ri qiymat: {e}")
    wb.close()

    if not rows:
        raise BadFile("Faylda ma'lumot qatori yo'q")
    return rows


def _copy(cur, table, cols, buf):
    buf.seek(0)
    cur.copy_from(buf, table, sep="\t", null=r"\N", columns=cols)


def load_fakt(source_file, rows, replace=False, conn=None):
    """Fakt savdo — bitta kunlik hisobot.

    Dublikat himoyasi SANA bo'yicha (fayl nomi bo'yicha emas — nomlar xilma-xil:
    `report (3).xlsx`, `5465132132151515.xlsx` ...).
    """
    dates = {r["date"] for r in rows}
    if len(dates) > 1:
        raise BadFile(
            f"Faylda {len(dates)} xil sana bor ({min(dates)} … {max(dates)}). "
            "Fakt savdo fayli bitta kunga tegishli bo'lishi kerak. "
            "Yakuniy savdo faylini yuklamoqchi bo'lsangiz — load_yakuniy() ishlating."
        )
    sale_date = next(iter(dates))

    own = conn is None
    c = conn or psycopg2.connect(DSN)
    try:
        cur = c.cursor()
        cur.execute("SELECT count(*), min(source_file) FROM fakt_savdo WHERE sale_date = %s",
                    (sale_date,))
        n, old_file = cur.fetchone()

        if n and not replace:
            return {"status": "skipped", "sale_date": sale_date, "rows": 0,
                    "reason": f"{sale_date} allaqachon yuklangan (fayl: {old_file})"}
        if n and replace:
            cur.execute("DELETE FROM fakt_savdo WHERE sale_date = %s", (sale_date,))

        buf = io.StringIO()
        for r in rows:
            buf.write("\t".join([
                r["date"].isoformat(), esc(r["agent"]), esc(r["orderer"]),
                esc(r["courier"]), esc(r["zone"]), esc(r["shop_type"]),
                esc(r["shop_name"]), esc(r["shop_no"]), esc(r["product_type"]),
                esc(r["product"]), esc(r["order_no"]), esc(r["pay_type"]),
                (f"{r['discount_pct']:.2f}" if r["discount_pct"] is not None else r"\N"),
                str(int(r["qty"])), f"{r['amount']:.2f}", source_file,
            ]) + "\n")
        _copy(cur, "fakt_savdo", FAKT_COLS, buf)
        c.commit()

        return {"status": "replaced" if n else "loaded", "sale_date": sale_date,
                "rows": len(rows), "qty": sum(r["qty"] for r in rows),
                "amount": sum(r["amount"] for r in rows)}
    finally:
        if own:
            c.close()


def load_yakuniy(source_file, rows, replace=False, conn=None):
    """Yakuniy savdo — haftalik fayl, bir nechta kun bo'lishi mumkin."""
    week = os.path.splitext(source_file)[0]
    own = conn is None
    c = conn or psycopg2.connect(DSN)
    try:
        cur = c.cursor()
        cur.execute("SELECT count(*) FROM yakuniy_savdo WHERE source_file = %s",
                    (source_file,))
        n = cur.fetchone()[0]
        if n and not replace:
            return {"status": "skipped", "rows": 0,
                    "reason": f"{source_file} allaqachon yuklangan"}
        if n and replace:
            cur.execute("DELETE FROM yakuniy_savdo WHERE source_file = %s", (source_file,))

        buf = io.StringIO()
        for r in rows:
            buf.write("\t".join([
                r["date"].isoformat(), week, esc(r["agent"]), esc(r["orderer"]),
                esc(r["courier"]), esc(r["zone"]), esc(r["shop_type"]),
                esc(r["shop_name"]), esc(r["shop_no"]), esc(r["product_type"]),
                esc(r["product"]), esc(r["order_no"]), esc(r["pay_type"]),
                (f"{r['discount_pct']:.2f}" if r["discount_pct"] is not None else r"\N"),
                f"{r['qty']:.2f}", f"{r['amount']:.2f}", source_file,
            ]) + "\n")
        _copy(cur, "yakuniy_savdo", YAK_COLS, buf)
        c.commit()

        ds = sorted({r["date"] for r in rows})
        return {"status": "replaced" if n else "loaded", "rows": len(rows),
                "kunlar": len(ds), "dan": ds[0], "gacha": ds[-1],
                "qty": sum(r["qty"] for r in rows)}
    finally:
        if own:
            c.close()


def refresh_views(conn=None):
    """Agregatlarni yangilash. PROGNOZNI QAYTA HISOBLAMAYDI."""
    own = conn is None
    c = conn or psycopg2.connect(DSN)
    try:
        cur = c.cursor()
        for v in ("mv_talab", "mv_sotilgan", "mv_talab_zich", "mv_mavsum", "mv_dokon_ulush"):
            cur.execute(f"REFRESH MATERIALIZED VIEW {v}")
        cur.execute("SELECT fn_products_sync()")
        cur.execute("ANALYZE fakt_savdo")
        c.commit()
    finally:
        if own:
            c.close()
